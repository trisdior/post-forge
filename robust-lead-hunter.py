#!/usr/bin/env python3
"""
Robust Lead Hunter - Focus on homeowner sections
"""

import json
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import quote
from urllib.error import URLError

# Configuration
SEEN_FILE = Path(r'C:\Users\trisd\clawd\data\craigslist-seen.json')
REPORT_FILE = Path(r'C:\Users\trisd\clawd\data\lead-hunt-report.json')
TRACKER_PATH = Path(r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx')

# Focus on sections where homeowners ask for services
SEARCHES = {
    'hss': ['painter', 'help', 'repair', 'bathroom', 'kitchen', 'need'],  # Services wanted
    'hsh': ['painter', 'help', 'repair', 'bathroom', 'kitchen', 'need'],  # Household services
}

# Red flags - contractor/business ads
RED_FLAGS = [
    r'\bhiring\b',
    r'\bposition\b',
    r'\bemployment\b',
    r'\bapplication\b',
    r'\bresume\b',
    r'\bpay\s*scale\b',
    r'\bsalary\b',
    r'\bcompensation\b',
    r'\bbenefits\b',
    r'\b(?:full[- ]?time|part[- ]?time|temporary)\b',
    r'\bcontractor\s+(?:needed|wanted|sought)\b',
    r'\bseek(?:ing)?\s+(?:professional|experienced)\b',
    r'\bwe are looking\b',
    r'\brequiring\b',
]

# Homeowner indicators - strong signals
HOMEOWNER_PATTERNS = [
    (r'\b(?:I need|I\'m looking|I want|I have|we need|my (?:home|house|place|apt))\b', 'personal need'),
    (r'\b(?:problem|damage|broken|leak|cracked|needs (?:repair|fixing)|help with)\b', 'specific problem'),
    (r'\b(?:bathroom|kitchen|bedroom|living|flooring|paint|drywall|roof|window)\b', 'room/area'),
]

def load_seen():
    """Load already-seen post IDs"""
    if SEEN_FILE.exists():
        try:
            with open(SEEN_FILE) as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except:
            return {}
    return {}

def save_seen(seen):
    """Save seen post IDs"""
    SEEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SEEN_FILE, 'w') as f:
        json.dump(seen, f, indent=2)

def fetch_search_listings(section, query):
    """Fetch all listing URLs for a search"""
    url = f"https://chicago.craigslist.org/search/{section}?query={quote(query)}"
    
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        req = Request(url, headers=headers)
        with urlopen(req, timeout=20) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Extract all /d/ URLs (listings)
        links = re.findall(r'(https://chicago\.craigslist\.org/[^"]*?/d/[^"]*?\.html)', html)
        
        listings = []
        seen_ids = set()
        for link in links:
            id_match = re.search(r'/(\d+)\.html', link)
            if id_match:
                post_id = id_match.group(1)
                if post_id not in seen_ids:
                    listings.append({'postId': post_id, 'postUrl': link, 'section': section})
                    seen_ids.add(post_id)
        
        return listings
    except Exception as e:
        print(f"    Error: {e}")
        return []

def extract_post_text(url):
    """Extract all text from a post (title + body)"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        req = Request(url, headers=headers)
        with urlopen(req, timeout=20) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Extract body text
        body_match = re.search(r'<section id="postingbody"[^>]*>(.+?)</section>', html, re.DOTALL)
        if body_match:
            text = body_match.group(1)
        else:
            text = ''
        
        # Clean HTML
        text = re.sub(r'<[^>]+>', ' ', text)
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        
        return text
    except:
        return None

def is_homeowner_post(text):
    """Check if text is from a homeowner asking for help"""
    if not text:
        return False, "No text"
    
    text_lower = text.lower()
    
    # Check for RED FLAGS (contractor/hiring ads)
    for pattern in RED_FLAGS:
        if re.search(pattern, text_lower):
            return False, "Contractor/hiring ad detected"
    
    # Count HOMEOWNER INDICATORS
    indicator_count = 0
    for pattern, label in HOMEOWNER_PATTERNS:
        if re.search(pattern, text_lower):
            indicator_count += 1
    
    if indicator_count >= 2:
        return True, f"{indicator_count} homeowner signals"
    else:
        return False, f"Only {indicator_count} homeowner signal(s)"

def log_to_tracker(lead):
    """Log lead to Excel tracker"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(TRACKER_PATH)
        ws = wb[wb.sheetnames[0]]
        
        # Find next empty row
        for row in range(3, 1000):
            if ws[f'B{row}'].value is None:
                ws[f'B{row}'] = datetime.now()
                ws[f'C{row}'] = 'Craigslist'
                ws[f'D{row}'] = 'Home Service'
                ws[f'H{row}'] = 'Chicago'
                ws[f'I{row}'] = lead['title'][:100]
                ws[f'K{row}'] = 'Hot'
                ws[f'L{row}'] = 'New'
                ws[f'P{row}'] = lead['url']
                break
        
        wb.save(TRACKER_PATH)
        return True
    except:
        return False

def main():
    seen = load_seen()
    valid_leads = []
    total_checked = 0
    
    print("\n" + "="*70)
    print("CRAIGSLIST HOMEOWNER LEAD HUNT")
    print("="*70)
    print(f"Time: {datetime.now().strftime('%I:%M %p Chicago')}")
    print()
    
    # Search homeowner-focused sections
    for section in SEARCHES:
        for query in SEARCHES[section]:
            print(f"[*] Searching [{section}] '{query}'...", end=" ")
            
            listings = fetch_search_listings(section, query)
            print(f"Found {len(listings)} posts")
            
            for listing in listings:
                post_id = listing['postId']
                
                # Skip if already processed
                if post_id in seen:
                    continue
                
                # Get post text
                post_text = extract_post_text(listing['postUrl'])
                if not post_text:
                    continue
                
                total_checked += 1
                
                # Validate
                is_valid, reason = is_homeowner_post(post_text)
                
                if is_valid:
                    # Extract title (first line or first sentence)
                    title_match = re.match(r'^(.+?)(?:\n|\. )', post_text)
                    title = title_match.group(1).strip() if title_match else post_text[:60]
                    
                    # Clean title for display (remove special chars)
                    display_title = title.encode('utf-8', errors='ignore').decode('utf-8')[:60]
                    
                    lead = {
                        'postId': post_id,
                        'title': title,
                        'url': listing['postUrl'],
                        'section': section,
                        'snippet': post_text[:200]
                    }
                    
                    print(f"    [+] VALID: {display_title}")
                    
                    valid_leads.append(lead)
                    seen[post_id] = datetime.now().isoformat()
                    log_to_tracker(lead)
                
                time.sleep(0.3)
            
            time.sleep(0.5)
    
    # Save progress
    save_seen(seen)
    
    # Deduplicate
    unique = {lead['postId']: lead for lead in valid_leads}
    final_leads = list(unique.values())
    
    # Quality score
    quality = round((len(final_leads) / max(total_checked, 1) * 100), 1)
    
    # Report
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'postsChecked': total_checked,
            'validLeads': len(final_leads),
            'qualityScore': quality,
        },
        'leads': final_leads
    }
    
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, indent=2)
    
    # Print results
    print("\n" + "="*70)
    print("SCAN RESULTS")
    print("="*70)
    print(f"Posts Checked: {total_checked}")
    print(f"Valid Homeowner Leads: {len(final_leads)}")
    print(f"Quality Score: {quality:.1f}%")
    
    if final_leads:
        print(f"\nTOP LEADS:")
        for i, lead in enumerate(final_leads[:10], 1):
            print(f"\n  {i}. {lead['title']}")
            print(f"     Section: {lead['section']} | Post ID: {lead['postId']}")
            print(f"     URL: {lead['url']}")
    
    print("\n" + "="*70)
    print(f"Report: {REPORT_FILE}")
    print("="*70 + "\n")

if __name__ == '__main__':
    main()
