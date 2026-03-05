#!/usr/bin/env python3
"""
Improved Lead Hunter for Valencia Construction
Searches proper Craigslist sections for homeowner requests
"""

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import quote
from urllib.error import URLError, HTTPError

# Configuration
SEEN_FILE = Path(r'C:\Users\trisd\clawd\data\craigslist-seen.json')
REPORT_FILE = Path(r'C:\Users\trisd\clawd\data\lead-hunt-report.json')
TRACKER_PATH = Path(r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx')

# Target sections and search terms
SEARCHES = {
    'lbg': [  # Labor gigs
        'painter needed',
        'help with kitchen',
        'bathroom repair needed',
        'drywall repair',
        'flooring install',
        'roof repair',
        'water damage'
    ],
    'dmg': [  # Domestic gigs (services wanted)
        'painter needed',
        'help with kitchen',
        'bathroom repair needed',
    ]
}

# Red flags - skip these
RED_FLAGS = [
    r'\bwe offer\b',
    r'\bwe specialize\b',
    r'\bwe do\b',
    r'\blicensed and insured\b',
    r'\byears of experience\b',
    r'\bfamily owned\b',
    r'\bcall us\b',
    r'\bcontact us\b',
    r'\w+\s+(?:llc|inc|co\.?|corp|construction|company)\b',
    r'\b(?:W2|1099)\b',
    r'\b(?:hiring|employment|job posting|position)\b'
]

# Homeowner indicators
HOMEOWNER_PATTERNS = [
    (r'\b(?:I|we|my|our|our home|our house|my house|my place)\b', 'personal pronouns'),
    (r'\b(?:need help|need someone|looking for|seeking|urgent|asap)\b', 'requesting help'),
    (r'\b(?:damage|repair|fix|broken|leak|cracked|water damage|drywall)\b', 'specific issue'),
    (r'\b(?:this weekend|next week|soon|flexible|available)\b', 'timeline'),
]

def load_seen():
    """Load already-seen post IDs"""
    if SEEN_FILE.exists():
        try:
            with open(SEEN_FILE) as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_seen(seen):
    """Save seen post IDs"""
    SEEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SEEN_FILE, 'w') as f:
        json.dump(seen, f, indent=2)

def search_craigslist_section(section, term, max_attempts=3):
    """Search Craigslist for a term in a specific section"""
    encoded = quote(term)
    url = f"https://chicago.craigslist.org/search/{section}?query={encoded}"
    
    print(f"[*] Searching [{section}]: '{term}'")
    results = []
    
    for attempt in range(max_attempts):
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            req = Request(url, headers=headers)
            with urlopen(req, timeout=30) as response:
                html = response.read().decode('utf-8', errors='ignore')
            
            # Extract listing links - new Craigslist URL format
            pattern = r'<a href="(https://chicago\.craigslist\.org/[^"]*?/(\d+))"[^>]*>([^<]+)</a>'
            matches = re.finditer(pattern, html)
            
            for m in matches:
                post_url, post_id, title = m.groups()
                # Skip non-listing URLs
                if '/d/' not in post_url or post_id.isdigit():
                    results.append({
                        'postUrl': post_url,
                        'postId': post_id,
                        'title': title.strip()[:100],
                        'section': section
                    })
            
            print(f"    Found {len(results)} results")
            return results
            
        except Exception as e:
            print(f"    Error (attempt {attempt+1}): {e}")
            if attempt < max_attempts - 1:
                time.sleep(2)
            else:
                return results
    
    return results

def validate_post(post_url, post_id, title, section):
    """Fetch and validate a post"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        req = Request(post_url, headers=headers)
        with urlopen(req, timeout=30) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Extract post body
        body_match = re.search(
            r'<section id="postingbody"[^>]*>(.+?)</section>',
            html,
            re.DOTALL
        )
        
        if body_match:
            post_text = body_match.group(1)
        else:
            post_text = title
        
        # Clean HTML
        post_text = re.sub(r'<[^>]+>', ' ', post_text)
        post_text = re.sub(r'\s+', ' ', post_text)
        post_text = re.sub(r'&quot;', '"', post_text)
        post_text = re.sub(r'&amp;', '&', post_text)
        post_text = post_text.strip()
        
        full_text = f"{title} {post_text}"
        
        # Validation result
        validation = {
            'isValid': True,
            'score': 0,
            'reason': '',
            'indicators': [],
            'redFlags': []
        }
        
        # Check RED FLAGS
        for pattern in RED_FLAGS:
            if re.search(pattern, full_text, re.IGNORECASE):
                validation['isValid'] = False
                validation['redFlags'].append(pattern[:30])
        
        # Check HOMEOWNER INDICATORS
        indicator_count = 0
        for pattern, label in HOMEOWNER_PATTERNS:
            if re.search(pattern, full_text, re.IGNORECASE):
                validation['indicators'].append(label)
                indicator_count += 1
                validation['score'] += 25
        
        # Final decision
        if validation['isValid']:
            if indicator_count < 2:
                validation['isValid'] = False
                validation['reason'] = f"Not enough homeowner signals ({indicator_count} found)"
            else:
                validation['reason'] = "Valid homeowner lead"
        else:
            validation['reason'] = f"Red flags detected"
        
        return {
            'postId': post_id,
            'title': title,
            'url': post_url,
            'section': section,
            'textSnippet': post_text[:300],
            'validation': validation,
            'valid': validation['isValid']
        }
    
    except Exception as e:
        print(f"    Error validating {post_id}: {e}")
        return None

def log_lead_to_tracker(lead):
    """Log a valid lead to the tracker spreadsheet"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(TRACKER_PATH)
        
        # Try to find the right sheet
        sheet_name = "Craigslist Leads"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]  # Use first sheet if Craigslist sheet doesn't exist
        
        ws = wb[sheet_name]
        
        # Find next empty row
        next_row = 3
        for row in range(3, 503):
            if ws[f'B{row}'].value is None:
                next_row = row
                break
        
        # Write lead data
        ws[f'B{next_row}'] = datetime.now()
        ws[f'B{next_row}'].number_format = 'MM/DD/YYYY'
        ws[f'C{next_row}'] = 'Craigslist'
        ws[f'D{next_row}'] = 'Home Remodel'
        ws[f'E{next_row}'] = ''  # Client name unknown
        ws[f'F{next_row}'] = ''  # Phone unknown
        ws[f'G{next_row}'] = ''  # Email unknown
        ws[f'H{next_row}'] = 'Chicago'
        ws[f'I{next_row}'] = lead['title']
        ws[f'K{next_row}'] = 'Hot'  # Default priority
        ws[f'L{next_row}'] = 'New'
        ws[f'O{next_row}'] = 'No'
        ws[f'P{next_row}'] = lead['url']
        
        wb.save(TRACKER_PATH)
        print(f"    Logged: {lead['title'][:40]}")
        return True
    except Exception as e:
        print(f"    Error logging to tracker: {e}")
        return False

def main():
    seen = load_seen()
    valid_leads = []
    rejected_leads = []
    total_checked = 0
    
    # Search all sections and terms
    for section, terms in SEARCHES.items():
        for term in terms:
            results = search_craigslist_section(section, term)
            time.sleep(1)
            
            for result in results:
                post_id = result['postId']
                
                # Skip if already seen
                if post_id in seen:
                    print(f"  [~] SKIP (seen): {result['title'][:60]}")
                    continue
                
                total_checked += 1
                validation = validate_post(
                    result['postUrl'],
                    post_id,
                    result['title'],
                    result['section']
                )
                
                if validation:
                    if validation['valid']:
                        print(f"  [+] VALID: {result['title'][:60]}")
                        print(f"      Indicators: {', '.join(validation['validation']['indicators'])}")
                        valid_leads.append(validation)
                        seen[post_id] = datetime.now().isoformat()
                        
                        # Log to tracker
                        log_lead_to_tracker(validation)
                    else:
                        print(f"  [-] REJECTED: {result['title'][:60]}")
                        print(f"      Reason: {validation['validation']['reason']}")
                        rejected_leads.append(validation)
                
                time.sleep(0.5)
            
            time.sleep(1)
    
    # Save seen posts
    save_seen(seen)
    
    # Deduplicate valid leads (in case same lead found in multiple sections)
    unique_leads = {}
    for lead in valid_leads:
        key = (lead['title'].lower(), lead['postId'])
        if key not in unique_leads:
            unique_leads[key] = lead
    
    valid_leads = list(unique_leads.values())
    
    # Generate report
    quality_score = (len(valid_leads) / total_checked * 100) if total_checked > 0 else 0
    
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'totalSearched': total_checked,
            'validLeadsFound': len(valid_leads),
            'rejectedLeads': len(rejected_leads),
            'qualityScore': round(quality_score, 1),
            'deduped': len(valid_leads)
        },
        'validLeads': valid_leads,
        'rejectedLeads': rejected_leads[:10]  # Top 10 for audit
    }
    
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, indent=2)
    
    # Print summary
    print("\n" + "="*60)
    print("CRAIGSLIST LEAD HUNT SUMMARY")
    print("="*60)
    print(f"Total Posts Checked: {total_checked}")
    print(f"Valid Homeowner Leads Found: {len(valid_leads)}")
    print(f"Rejected Posts: {len(rejected_leads)}")
    print(f"Quality Score: {quality_score:.1f}%")
    print(f"Time: {datetime.now().strftime('%I:%M %p')} Chicago")
    
    if valid_leads:
        print("\nVALID HOMEOWNER LEADS:")
        for i, lead in enumerate(valid_leads, 1):
            print(f"  {i}. {lead['title']}")
            print(f"     Section: {lead['section']} | URL: {lead['url'][:60]}...")
    
    print(f"\nReport saved to: {REPORT_FILE}")
    print("="*60)
    
    return json.dumps(report, indent=2)

if __name__ == '__main__':
    result = main()
