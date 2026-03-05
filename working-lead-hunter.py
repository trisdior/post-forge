#!/usr/bin/env python3
"""
Working Lead Hunter - Uses proven search sections
Searches Craigslist for homeowner service requests
"""

import json
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import quote

# Configuration
SEEN_FILE = Path(r'C:\Users\trisd\clawd\data\craigslist-seen.json')
REPORT_FILE = Path(r'C:\Users\trisd\clawd\data\lead-hunt-report.json')
TRACKER_PATH = Path(r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx')

# Sections that return results + best terms for homeowner requests
SEARCHES = {
    'lbg': ['painter', 'help', 'need', 'repair'],      # Labor gigs
    'hss': ['painter', 'help', 'need', 'repair'],      # Services wanted
    'hsh': ['painter', 'help', 'need', 'repair'],      # Household services
}

# Red flags - skip these (contractor ads)
RED_FLAGS = [
    r'\bwe offer\b',
    r'\bwe specialize\b',
    r'\bwe do\b',
    r'\blicensed and insured\b',
    r'\byears of experience\b',
    r'\bfamily owned\b',
    r'\bcall us\b',
    r'\bcontact us\b',
    r'\w+\s+(?:llc|inc|co\.?|corp|construction|company)\b',
    r'\b(?:W2|1099)\b',
    r'\b(?:hiring|employment|job posting|position)\b',
    r'\bservice\s+(?:provider|professional|expert)',
]

# Homeowner indicators
HOMEOWNER_PATTERNS = [
    (r'\b(?:I|we|my|our|our home|our house|my house|my place)\b', 'personal pronouns'),
    (r'\b(?:need (?:a|someone|help)|looking for|seeking|urgent|asap|need someone)\b', 'requesting help'),
    (r'\b(?:damage|repair|fix|broken|leak|cracked|water damage|drywall|paint|bathroom|kitchen)\b', 'specific project'),
    (r'\b(?:this weekend|next week|soon|flexible|available|can start|when available)\b', 'timeline'),
]

def load_seen():
    """Load already-seen post IDs"""
    if SEEN_FILE.exists():
        try:
            with open(SEEN_FILE) as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_seen(seen):
    """Save seen post IDs"""
    SEEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SEEN_FILE, 'w') as f:
        json.dump(seen, f, indent=2)

def fetch_listings(section, query):
    """Fetch listings from a Craigslist search"""
    url = f"https://chicago.craigslist.org/search/{section}?query={quote(query)}"
    
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        req = Request(url, headers=headers)
        with urlopen(req, timeout=20) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        listings = []
        
        # Find all listing links - improved pattern
        pattern = r'href="(https://chicago\.craigslist\.org/([a-z]{3})/[a-z]{3}/d/[^"]+?)">([^<]*)</a>'
        for match in re.finditer(pattern, html):
            url_full, area, title = match.groups()
            # Extract numeric ID from URL
            id_match = re.search(r'/(\d+)(?:\.|$)', url_full)
            if id_match:
                post_id = id_match.group(1)
                listings.append({
                    'postId': post_id,
                    'postUrl': url_full,
                    'title': title.strip()[:100],
                    'area': area,
                    'section': section,
                    'query': query
                })
        
        return listings
    except Exception as e:
        print(f"    Error fetching {section}/{query}: {e}")
        return []

def fetch_post_body(url):
    """Fetch the full post body"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        req = Request(url, headers=headers)
        with urlopen(req, timeout=20) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Extract post body
        body_match = re.search(
            r'<section id="postingbody"[^>]*>(.+?)</section>',
            html,
            re.DOTALL
        )
        
        if body_match:
            text = body_match.group(1)
        else:
            return None
        
        # Clean HTML
        text = re.sub(r'<[^>]+>', ' ', text)
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'&quot;', '"', text)
        text = re.sub(r'&amp;', '&', text)
        text = re.sub(r'&lt;', '<', text)
        text = re.sub(r'&gt;', '>', text)
        text = text.strip()
        
        return text
    except:
        return None

def validate_listing(post_id, title, url, section):
    """Validate if listing is from a homeowner asking for help"""
    
    # Fetch full post body
    post_body = fetch_post_body(url)
    if not post_body:
        return None
    
    full_text = f"{title} {post_body}"
    
    # Check RED FLAGS
    red_flag_matches = []
    for pattern in RED_FLAGS:
        if re.search(pattern, full_text, re.IGNORECASE):
            red_flag_matches.append(pattern[:25])
    
    if red_flag_matches:
        return {
            'postId': post_id,
            'title': title,
            'url': url,
            'section': section,
            'valid': False,
            'reason': 'Red flags detected',
            'redFlags': red_flag_matches[:2]
        }
    
    # Check HOMEOWNER INDICATORS
    indicators = []
    for pattern, label in HOMEOWNER_PATTERNS:
        if re.search(pattern, full_text, re.IGNORECASE):
            indicators.append(label)
    
    # Decision logic
    is_valid = len(indicators) >= 2  # Need at least 2 indicators
    
    return {
        'postId': post_id,
        'title': title,
        'url': url,
        'section': section,
        'valid': is_valid,
        'reason': f"{len(indicators)} homeowner signals" if is_valid else f"Only {len(indicators)} signal(s)",
        'indicators': indicators,
        'textSnippet': post_body[:200]
    }

def log_to_tracker(lead):
    """Add lead to Excel tracker"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(TRACKER_PATH)
        ws = wb[wb.sheetnames[0]]  # Use first sheet
        
        # Find next empty row
        next_row = 3
        for row in range(3, 503):
            if ws[f'B{row}'].value is None:
                next_row = row
                break
        
        # Write lead
        ws[f'B{next_row}'] = datetime.now()
        ws[f'B{next_row}'].number_format = 'MM/DD/YYYY'
        ws[f'C{next_row}'] = 'Craigslist'
        ws[f'D{next_row}'] = 'Home Remodel'
        ws[f'E{next_row}'] = ''
        ws[f'F{next_row}'] = ''
        ws[f'G{next_row}'] = ''
        ws[f'H{next_row}'] = 'Chicago'
        ws[f'I{next_row}'] = lead['title']
        ws[f'K{next_row}'] = 'Hot'
        ws[f'L{next_row}'] = 'New'
        ws[f'O{next_row}'] = 'No'
        ws[f'P{next_row}'] = lead['url']
        
        wb.save(TRACKER_PATH)
        return True
    except Exception as e:
        print(f"  Error logging: {e}")
        return False

def generate_followups(leads_count):
    """Generate follow-up templates"""
    if leads_count == 0:
        return []
    
    followups = [
        {
            'template': 'FIRST_CONTACT',
            'message': "Hi! I saw your post on Craigslist. We're Valencia Construction, and we specialize in exactly what you're looking for. I'd love to help—can I come by for a free estimate?",
            'timing': 'Immediate'
        },
        {
            'template': 'NO_RESPONSE_24H',
            'message': "Just following up on my message about your project. Are you still looking for someone? I'm ready to help ASAP.",
            'timing': '24 hours'
        },
        {
            'template': 'NO_RESPONSE_48H',
            'message': "One more quick message—we have availability this week. Let me know if you'd like that free estimate.",
            'timing': '48 hours'
        }
    ]
    
    return followups

def main():
    seen = load_seen()
    valid_leads = []
    rejected_leads = []
    total_checked = 0
    
    print("\n" + "="*70)
    print("CRAIGSLIST HOMEOWNER LEAD SCANNER")
    print("="*70)
    print(f"Time: {datetime.now().strftime('%I:%M %p %Z')} | Chicago Market")
    print()
    
    # Search each section + term combo
    for section in SEARCHES:
        for query in SEARCHES[section]:
            print(f"[*] Searching [{section:3}] '{query}'...", end=" ")
            
            listings = fetch_listings(section, query)
            print(f"Found {len(listings)} listings")
            
            for listing in listings:
                post_id = listing['postId']
                
                # Skip if already seen
                if post_id in seen:
                    continue
                
                total_checked += 1
                
                # Validate the listing
                validation = validate_listing(
                    post_id,
                    listing['title'],
                    listing['postUrl'],
                    section
                )
                
                if validation:
                    if validation['valid']:
                        print(f"    ✓ VALID: {listing['title'][:55]}")
                        print(f"          {', '.join(validation.get('indicators', []))}")
                        valid_leads.append(validation)
                        seen[post_id] = datetime.now().isoformat()
                        log_to_tracker(validation)
                    else:
                        rejected_leads.append(validation)
                        print(f"    ✗ SKIP: {listing['title'][:55]}")
                
                time.sleep(0.3)
            
            time.sleep(0.5)
    
    # Save and deduplicate
    save_seen(seen)
    
    # Remove duplicates
    unique_leads = {}
    for lead in valid_leads:
        key = (lead['postId'], lead['title'].lower())
        if key not in unique_leads:
            unique_leads[key] = lead
    
    valid_leads = list(unique_leads.values())
    
    # Quality score
    quality = round((len(valid_leads) / total_checked * 100), 1) if total_checked > 0 else 0
    
    # Generate report
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'totalSearched': total_checked,
            'validLeads': len(valid_leads),
            'rejectedLeads': len(rejected_leads),
            'qualityScore': quality,
        },
        'validLeads': valid_leads,
        'followUpSequence': generate_followups(len(valid_leads))
    }
    
    # Save report
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, indent=2)
    
    # Print summary
    print("\n" + "="*70)
    print("SCAN COMPLETE")
    print("="*70)
    print(f"Posts Checked: {total_checked}")
    print(f"Valid Homeowner Leads: {len(valid_leads)}")
    print(f"Quality Score: {quality}%")
    print()
    
    if valid_leads:
        print("LEADS FOUND:")
        for i, lead in enumerate(valid_leads, 1):
            print(f"\n  {i}. {lead['title']}")
            print(f"     Section: [{lead['section']}] | Craigslist Post ID: {lead['postId']}")
            print(f"     URL: {lead['url'][:65]}...")
            print(f"     Snippet: {lead.get('textSnippet', '')[:80]}...")
    
    if valid_leads:
        print("\n" + "-"*70)
        print("FOLLOW-UP SEQUENCE:")
        for followup in generate_followups(len(valid_leads)):
            print(f"\n  [{followup['timing']}] {followup['template']}")
            print(f"  Message: {followup['message']}")
    
    print("\n" + "="*70)
    print(f"Report: {REPORT_FILE}")
    print("="*70 + "\n")
    
    return report

if __name__ == '__main__':
    main()
