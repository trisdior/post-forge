#!/usr/bin/env python3
"""
Final Working Lead Hunter for Valencia Construction
Fetches individual posts and validates them
"""

import json
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import quote

# Configuration
SEEN_FILE = Path(r'C:\Users\trisd\clawd\data\craigslist-seen.json')
REPORT_FILE = Path(r'C:\Users\trisd\clawd\data\lead-hunt-report.json')
TRACKER_PATH = Path(r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx')

# Sections + terms that return results
SEARCHES = {
    'lbg': ['painter', 'help', 'need', 'repair'],
    'hss': ['painter', 'help', 'need', 'repair'],
    'hsh': ['painter', 'help', 'need', 'repair'],
}

# Red flags - contractor/hiring ads
RED_FLAGS = [
    r'\bwe (?:offer|specialize|do|provide|can)\b',
    r'\blicensed and insured\b',
    r'\byears of experience\b',
    r'\bfamily owned\b',
    r'\bcall us\b',
    r'\bcontact us\b',
    r'\b(?:hiring|employment|job|position|opportunity|now hiring)\b',
    r'\b(?:apprentice|trainee|assistant|work with us)\b',
    r'\bW-?2|1099\b',
    r'\bcompany\b|\bllc\b|\binc\b|\bcorp\b',
]

# Homeowner indicators
HOMEOWNER_PATTERNS = [
    (r'\b(?:I|we|my|our|mine|ours)\b', 'personal pronouns'),
    (r'\b(?:need (?:a |someone|help)|looking for|seeking|help with|can you)\b', 'request for help'),
    (r'\b(?:damage|repair|fix|broken|leak|cracked|paint|bathroom|kitchen|flooring|drywall)\b', 'project type'),
    (r'\b(?:this week|next week|asap|urgent|soon|flexible|available)\b', 'timeline'),
]

def load_seen():
    """Load already-seen post IDs"""
    if SEEN_FILE.exists():
        try:
            with open(SEEN_FILE) as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except:
            return {}
    return {}

def save_seen(seen):
    """Save seen post IDs"""
    SEEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SEEN_FILE, 'w') as f:
        json.dump(seen, f, indent=2)

def fetch_search_results(section, query):
    """Fetch listing URLs from a search result page"""
    url = f"https://chicago.craigslist.org/search/{section}?query={quote(query)}"
    
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        req = Request(url, headers=headers)
        with urlopen(req, timeout=20) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Find all listing links (they're complete URLs with .html)
        links = re.findall(r'(https://chicago\.craigslist\.org/[^"]*?/d/[^"]*?\.html)', html)
        
        # Extract post IDs (numeric part)
        listings = []
        for link in links:
            id_match = re.search(r'/(\d+)\.html', link)
            if id_match:
                post_id = id_match.group(1)
                listings.append({'postId': post_id, 'postUrl': link, 'section': section, 'query': query})
        
        return listings
    except Exception as e:
        print(f"    Error: {e}")
        return []

def fetch_post(url):
    """Fetch a single post and extract title + body"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        req = Request(url, headers=headers)
        with urlopen(req, timeout=20) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Extract title
        title_match = re.search(r'<span class="postingtitle"[^>]*><span class="postingtitletext"[^>]*>([^<]+)</span>', html)
        title = title_match.group(1).strip() if title_match else 'Unknown Title'
        
        # Clean title (remove post ID if present)
        title = re.sub(r'\s*pic\s*$', '', title, flags=re.IGNORECASE).strip()
        
        # Extract post body
        body_match = re.search(r'<section id="postingbody"[^>]*>(.+?)</section>', html, re.DOTALL)
        if body_match:
            body = body_match.group(1)
            # Clean HTML tags
            body = re.sub(r'<[^>]+>', ' ', body)
            body = re.sub(r'\s+', ' ', body)
            body = re.sub(r'&quot;|&apos;', '"', body)
            body = re.sub(r'&amp;', '&', body)
            body = re.sub(r'&lt;', '<', body)
            body = re.sub(r'&gt;', '>', body)
            body = body.strip()
        else:
            body = ''
        
        return {'title': title, 'body': body}
    except:
        return None

def validate_post(post_id, title, body, url, section):
    """Validate if this is a homeowner asking for help"""
    
    full_text = f"{title} {body}".lower()
    
    # Check for red flags
    red_flag_count = 0
    for pattern in RED_FLAGS:
        if re.search(pattern, full_text, re.IGNORECASE):
            red_flag_count += 1
    
    if red_flag_count > 0:
        return {'valid': False, 'reason': 'Contractor/hiring ad detected'}
    
    # Check for homeowner indicators
    indicator_matches = []
    for pattern, label in HOMEOWNER_PATTERNS:
        if re.search(pattern, full_text, re.IGNORECASE):
            indicator_matches.append(label)
    
    # Valid if at least 2 indicators
    is_valid = len(indicator_matches) >= 2
    
    return {
        'valid': is_valid,
        'indicators': indicator_matches,
        'reason': f"{len(indicator_matches)} homeowner signals" if is_valid else f"Only {len(indicator_matches)} signal"
    }

def log_to_tracker(lead):
    """Add to Excel tracker"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(TRACKER_PATH)
        ws = wb[wb.sheetnames[0]]
        
        # Find next empty row
        next_row = 3
        for row in range(3, 1000):
            if ws[f'B{row}'].value is None:
                next_row = row
                break
        
        ws[f'B{next_row}'] = datetime.now()
        ws[f'C{next_row}'] = 'Craigslist'
        ws[f'D{next_row}'] = 'Home Remodel'
        ws[f'H{next_row}'] = 'Chicago'
        ws[f'I{next_row}'] = lead['title'][:100]
        ws[f'K{next_row}'] = 'Hot'
        ws[f'L{next_row}'] = 'New'
        ws[f'P{next_row}'] = lead['url']
        
        wb.save(TRACKER_PATH)
        return True
    except:
        return False

def main():
    seen = load_seen()
    valid_leads = []
    rejected_leads = []
    total_checked = 0
    
    print("\n" + "="*70)
    print("CRAIGSLIST HOMEOWNER LEAD SCANNER - FINAL")
    print("="*70)
    print(f"Time: {datetime.now().strftime('%I:%M %p')} | Scanning for homeowner requests")
    print()
    
    # Search each section
    for section in SEARCHES:
        for query in SEARCHES[section]:
            print(f"[*] Searching [{section}] for '{query}'...", end=" ")
            
            listings = fetch_search_results(section, query)
            print(f"Found {len(listings)}")
            
            for listing in listings:
                post_id = listing['postId']
                
                # Skip if already seen
                if post_id in seen:
                    continue
                
                # Fetch the full post
                post_data = fetch_post(listing['postUrl'])
                if not post_data:
                    continue
                
                total_checked += 1
                
                # Validate
                validation = validate_post(
                    post_id,
                    post_data['title'],
                    post_data['body'],
                    listing['postUrl'],
                    section
                )
                
                if validation['valid']:
                    lead = {
                        'postId': post_id,
                        'title': post_data['title'],
                        'url': listing['postUrl'],
                        'section': section,
                        'indicators': validation['indicators'],
                        'snippet': post_data['body'][:150]
                    }
                    
                    print(f"    [+] VALID: {post_data['title'][:55]}")
                    print(f"               {', '.join(validation['indicators'])}")
                    
                    valid_leads.append(lead)
                    seen[post_id] = datetime.now().isoformat()
                    log_to_tracker(lead)
                
                # Rate limiting
                time.sleep(0.5)
            
            time.sleep(1)
    
    # Save seen posts
    save_seen(seen)
    
    # Deduplicate
    unique_keys = set()
    deduped_leads = []
    for lead in valid_leads:
        key = lead['postId']
        if key not in unique_keys:
            unique_keys.add(key)
            deduped_leads.append(lead)
    
    # Quality score
    quality = round((len(deduped_leads) / total_checked * 100), 1) if total_checked > 0 else 0
    
    # Report
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'postsChecked': total_checked,
            'validLeads': len(deduped_leads),
            'rejectedLeads': total_checked - len(deduped_leads),
            'qualityScore': quality,
        },
        'leads': deduped_leads[:50]  # Top 50
    }
    
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, indent=2)
    
    # Print summary
    print("\n" + "="*70)
    print("SCAN COMPLETE")
    print("="*70)
    print(f"Posts Checked: {total_checked}")
    print(f"[+] Valid Homeowner Leads: {len(deduped_leads)}")
    print(f"Quality Score: {quality:.1f}%")
    
    if deduped_leads:
        print(f"\nTOP LEADS:")
        for i, lead in enumerate(deduped_leads[:10], 1):
            print(f"\n  {i}. {lead['title']}")
            print(f"     Post ID: {lead['postId']} | Section: {lead['section']}")
            print(f"     Indicators: {', '.join(lead['indicators'])}")
            print(f"     URL: {lead['url']}")
    
    print(f"\nReport: {REPORT_FILE}")
    print("="*70 + "\n")
    
    return report

if __name__ == '__main__':
    main()
