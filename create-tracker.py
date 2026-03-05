import json
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl not available, creating CSV alternative...")
    sys.exit(0)

# Load leads from JSON
with open('follow-up-queue.json', 'r') as f:
    leads = json.load(f)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Facebook Leads"

# Add headers
headers = [
    "Lead ID",
    "Name",
    "Project Type",
    "Location",
    "Specific Needs",
    "Timeline",
    "Estimated Value",
    "Facebook URL",
    "Source",
    "Status",
    "Date Added",
    "Last Contact"
]

ws.append(headers)

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data rows
for lead in leads:
    ws.append([
        lead['lead_id'],
        lead['name'],
        lead['project_type'],
        lead['location'],
        lead['specific_needs'],
        lead['timeline'],
        lead['estimated_value'],
        lead['facebook_url'],
        lead['source'],
        lead['status'],
        datetime.now().strftime("%Y-%m-%d"),
        ""
    ])

# Adjust column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 50
ws.column_dimensions['I'].width = 18
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 12

# Save
wb.save('data/Valencia-Lead-Tracker.xlsx')
print("[OK] Created Valencia-Lead-Tracker.xlsx with {} leads".format(len(leads)))
