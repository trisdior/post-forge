#!/usr/bin/env python3
"""
Daily Business Status Check — Runs every morning to audit Valencia systems
Reports: Hunter status, Lead Tracker updates, Karl queue, Blockers
"""

import json
import os
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook

DATA_DIR = Path(r"C:\Users\trisd\clawd\data")
TRACKER_PATH = DATA_DIR / "Valencia-Lead-Tracker.xlsx"
AGENT_QUEUE = DATA_DIR / "agent-queue.json"
REPORT_FILE = DATA_DIR / "daily-status-report.json"

def get_tracker_stats():
    """Extract stats from Lead Tracker."""
    if not TRACKER_PATH.exists():
        return None
    
    try:
        wb = load_workbook(TRACKER_PATH)
        ws = wb["Lead Tracker"]
        
        # Count leads by status
        new_leads = 0
        contacted = 0
        total_leads = 0
        last_lead_date = None
        
        for row in range(3, 505):  # Check up to 500 rows
            date_cell = ws[f"B{row}"].value
            status_cell = ws[f"L{row}"].value
            
            if date_cell:
                total_leads += 1
                last_lead_date = date_cell
                if status_cell == "New":
                    new_leads += 1
                elif status_cell == "Contacted":
                    contacted += 1
        
        return {
            "total": total_leads,
            "new": new_leads,
            "contacted": contacted,
            "last_lead_date": last_lead_date.isoformat() if last_lead_date else None
        }
    except Exception as e:
        return {"error": str(e)}

def get_karl_queue_stats():
    """Check what's pending Karl's approval."""
    if not AGENT_QUEUE.exists():
        return {"total": 0, "new": 0, "pending_approval": 0}
    
    try:
        with open(AGENT_QUEUE) as f:
            queue = json.load(f)
        
        new = [item for item in queue if item.get("status") == "new"]
        pending = [item for item in queue if item.get("status") == "pending-approval"]
        
        return {
            "total": len(queue),
            "new": len(new),
            "pending_approval": len(pending),
            "hot_leads": sum(1 for item in new if item.get("score", 0) >= 50)
        }
    except Exception as e:
        return {"error": str(e)}

def check_hunter_health():
    """Check if Hunter is running and producing leads."""
    report_file = DATA_DIR / "lead-hunt-report.json"
    
    if not report_file.exists():
        return {"status": "No report", "last_run": None}
    
    try:
        with open(report_file) as f:
            report = json.load(f)
        
        timestamp = report.get("timestamp")
        summary = report.get("summary", {})
        
        last_run = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
        time_since = datetime.utcnow() - last_run.replace(tzinfo=None)
        
        return {
            "status": "Running",
            "last_run": timestamp,
            "hours_ago": round(time_since.total_seconds() / 3600, 1),
            "posts_checked": summary.get("totalSearched"),
            "valid_leads": summary.get("validLeads"),
            "quality": summary.get("qualityScore")
        }
    except Exception as e:
        return {"status": "Error", "error": str(e)}

def build_report():
    """Compile full business status report."""
    tracker = get_tracker_stats()
    karl = get_karl_queue_stats()
    hunter = check_hunter_health()
    
    report = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "tracker": tracker,
        "karl_queue": karl,
        "hunter": hunter,
        "alerts": []
    }
    
    # Generate alerts
    if hunter.get("hours_ago", 999) > 12:
        report["alerts"].append("[WARNING] Hunter hasn't run in 12+ hours")
    
    if karl.get("pending_approval", 0) > 0:
        report["alerts"].append(f"[APPROVAL PENDING] {karl['pending_approval']} lead(s) awaiting approval from Tris")
    
    if karl.get("hot_leads", 0) > 0:
        report["alerts"].append(f"[HOT LEADS] {karl['hot_leads']} hot lead(s) in queue (score >= 50)")
    
    if tracker and tracker.get("total", 0) == 0:
        report["alerts"].append("[ERROR] No leads in tracker yet")
    
    if tracker and tracker.get("new", 0) > 5:
        report["alerts"].append(f"[ACTION] {tracker['new']} new leads waiting for follow-up")
    
    return report

def print_report(report):
    """Pretty-print the status report."""
    print("\n" + "="*70)
    print(f"VALENCIA DAILY STATUS CHECK -- {datetime.now().strftime('%a, %b %d, %I:%M %p')}")
    print("="*70)
    
    # Hunter Status
    hunter = report.get("hunter", {})
    print(f"\n[HUNTER] Lead Scraper")
    if "error" in hunter:
        print(f"   ERROR: {hunter['error']}")
    else:
        status = hunter.get("status", "Unknown")
        if status == "Running":
            hours = hunter.get("hours_ago")
            print(f"   OK - Last run: {hours}h ago")
            print(f"   Scanned: {hunter.get('posts_checked')} posts")
            print(f"   Found: {hunter.get('valid_leads')} valid leads (quality: {hunter.get('quality')}%)")
        else:
            print(f"   WARNING - Status: {status}")
    
    # Lead Tracker
    tracker = report.get("tracker", {})
    print(f"\n[TRACKER] Lead Database")
    if "error" in tracker:
        print(f"   ERROR: {tracker['error']}")
    else:
        print(f"   Total: {tracker.get('total', 0)} leads")
        print(f"   New: {tracker.get('new', 0)}")
        print(f"   Contacted: {tracker.get('contacted', 0)}")
        if tracker.get("last_lead_date"):
            print(f"   Last: {tracker.get('last_lead_date')}")
    
    # Karl (Closer) Queue
    karl = report.get("karl_queue", {})
    print(f"\n[KARL] Reply Drafter")
    if "error" in karl:
        print(f"   ERROR: {karl['error']}")
    else:
        print(f"   Queued: {karl.get('total', 0)} leads")
        print(f"   New: {karl.get('new', 0)} (unprocessed)")
        print(f"   Pending Approval: {karl.get('pending_approval', 0)}")
        if karl.get("hot_leads", 0) > 0:
            print(f"   HOT Leads: {karl.get('hot_leads')}")
    
    # Alerts
    alerts = report.get("alerts", [])
    if alerts:
        print(f"\n[ALERTS]")
        for alert in alerts:
            print(f"   {alert}")
    else:
        print(f"\n[OK] No alerts -- all systems nominal")
    
    print("\n" + "="*70 + "\n")

def main():
    report = build_report()
    
    # Save report
    with open(REPORT_FILE, "w") as f:
        json.dump(report, f, indent=2)
    
    # Print to console
    print_report(report)
    
    # Return status for cron
    return 0 if not report.get("alerts") else 1

if __name__ == "__main__":
    exit(main())
