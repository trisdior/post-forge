"""
Integrate scanner leads into the Valencia Lead Tracker spreadsheet.
Reads all scanner JSON output files, scores each post for homeowner intent,
deduplicates, auto-categorizes, and writes qualified leads to the Lead Tracker
sheet so they appear in the Mission Control dashboard.

Scoring system: each post gets a homeowner_score from -100 to +100.
Only posts scoring >= 10 are logged. Marketplace (Angi/Thumbtack/HomeAdvisor)
is excluded — those are intelligence-only, not loggable leads.

Usage:
  python integrate-scanner-leads.py                  # Process all scanners
  python integrate-scanner-leads.py --source reddit  # Process one source
  python integrate-scanner-leads.py --dry-run        # Score without writing
"""

import json
import sys
import io
import os
import re
import uuid
import argparse
from datetime import datetime
from openpyxl import load_workbook

# Fix Unicode output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DATA_DIR = r"C:\Users\trisd\clawd\data"
TRACKER_PATH = os.path.join(DATA_DIR, "Valencia-Lead-Tracker.xlsx")
AGENT_QUEUE_PATH = os.path.join(DATA_DIR, "agent-queue.json")

# Scanner JSON files to process (marketplace EXCLUDED — intelligence only)
SCANNER_FILES = {
    "reddit": os.path.join(DATA_DIR, "reddit-leads-latest.json"),
    "nextdoor": os.path.join(DATA_DIR, "nextdoor-leads-latest.json"),
    "craigslist": os.path.join(DATA_DIR, "craigslist-latest.json"),
    "facebook": os.path.join(DATA_DIR, "facebook-leads-latest.json"),
}

# Minimum score to log a lead (scale: -100 to +100)
# Higher threshold for Craigslist (more spam), lower for Reddit/Facebook
SCORE_THRESHOLD = 15
SCORE_THRESHOLD_BY_SOURCE = {
    "craigslist": 30,  # High threshold — lots of contractor spam on CL
    "reddit": 10,      # Moderate — Reddit users are generally real homeowners
    "facebook": 10,    # Moderate — Facebook groups curate somewhat
    "nextdoor": 15,    # Moderate-high — Nextdoor has some business listings
}

# ── Homeowner Scoring System ─────────────────────────────────────────

# Negative signals: contractor/service provider ads
NEGATIVE_SIGNALS = [
    # HARD BLOCKS — instant disqualify
    # Phone number in title (contractors advertise their number)
    (r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", -100),
    # Multiple common services (contractors list everything)
    (r"(?i)(?:paint|plumb|electric|roof|floor|drywall|tile|hvac|carpent|concrete|demolition|gutter|siding|window|door|fence|deck|landscap|clean|junk|haul|mov).*(?:paint|plumb|electric|roof|floor|drywall|tile|hvac|carpent|concrete|demolition|gutter|siding|window|door|fence|deck|landscap|clean|junk|haul|mov).*(?:paint|plumb|electric|roof|floor|drywall|tile|hvac|carpent|concrete|demolition|gutter|siding|window|door|fence|deck|landscap|clean|junk|haul|mov)", -100),
    # LLC/Inc/Corp (business entity)
    (r"(?i)\bLLC\b|\bInc\b|\bCorp\b|\bCo\b|\bLtd\b|\&\s*Co\b", -100),
    
    # STRONG SIGNALS (contractor language)
    # ALL CAPS title (8+ consecutive uppercase words = ad style)
    (r"(?:^|\s)(?:[A-Z]{2,}\s+){3,}", -25),
    # Free estimate/quote (classic contractor language)
    (r"(?i)free\s+(estimate|quote|consult|inspection)", -35),
    # Licensed/insured/bonded (contractors brag about this)
    (r"(?i)licens\w*[,\s/&]*insur|insur\w*[,\s/&]*licens|(?:fully\s+)?(?:licensed|insured|bonded)", -40),
    # "We offer/provide/specialize/do/serve" (contractor pitch)
    (r"(?i)\bwe\s+(offer|provide|specializ|are\s+a|do|handle|service|serve|install|repair|build)", -35),
    # "Call us/contact us/reach us" (sales call-to-action)
    (r"(?i)(call|contact|reach|text|message|dm|email|call\s+now)\s+(us|today|now|for|a|the)", -35),
    # Years of experience (contractor credential)
    (r"(?i)\d+\+?\s*years?\s*(of\s+)?experience", -25),
    # Pricing in title ($299, $99, etc.)
    (r"\$\d{2,}", -35),
    # "Serving [area]" (contractor scope language)
    (r"(?i)serving\s+(chicago|illinois|northside|south\s*side|north\s*side|the\s+area|greater|all|any)", -25),
    # URL / website in post (contractor advertising)
    (r"(?i)www\.|\.com[/\s]|\.net[/\s]|\.org[/\s]|check\s+us\s+out", -30),
    # "Our/My services/team/company" (contractor self-promotion)
    (r"(?i)\b(our|my)\s+(services?|team|company|crew|staff|business|contractors?)", -30),
    # Quality/professional/expert (contractor marketing buzzwords)
    (r"(?i)(high.?quality|professional|expert|skilled|certified)\s+(work|service|craftsmanship|result|installation)", -25),
    # Affordable/competitive rates (price advertising)
    (r"(?i)(affordable|competitive|low.?cost|cheap|best\s+price)\s*(rates?|prices?|pricing)?", -25),
    # Hiring/job postings (not homeowner leads)
    (r"(?i)(hiring|now hiring|join our team|help wanted|looking for.*worker|need.*employee)", -100),
    # Asking for recommendations (not direct leads — they're asking others, not hiring themselves)
    (r"(?i)(recommend|recommendation|any good|anyone know|who do you.*|suggestions?|referrals?)\s+(contractor|plumber|painter|electrician|company|service|place|guy|someone)", -50),
    (r"(?i)(looking for.*recommend|can someone recommend|any.*recommend)", -40),
    # Service title pattern: "[Service] Service" or "NAME's [Service]"
    (r"(?i)\b(painting|plumbing|electric|handyman|construction|flooring|roofing)\s+services?\b", -35),
    (r"(?i)\b\w+'s\s+(painting|plumbing|electric|handyman|construction|flooring|roofing|renovation|repair)", -35),
    # "[Service] installs/repairs" = contractor listing
    (r"(?i)(installs?|repairs?|builds?|creates?|provides?)\s*(windows|doors|roofs|siding|gutters|decks|fences|flooring)", -30),
    # Emoji spam (contractors love emoji-heavy titles)
    (r"[\U0001f525\U0001f4a7\u26a1\U0001f3e0\U0001f6a8\U0001f4b0\U0001f947\u26ea\U0001f9f9\U0001f6c1\U0001f6aa\U0001f7e0\U0001f7e2\U0001f535]{3,}", -25),
    # "Better call [Name]" (famous contractor ad reference)
    (r"(?i)better\s+call\s+\w+", -50),
    # Repeated punctuation or spam characters (!!!, ???, ;;;)
    (r"([!?]{2,}|;{2,})", -20),
    # Random caps MiXeD (spam/nonsense indicator)
    (r"[a-z][A-Z]{2,}[a-z]", -15),
]

# Positive signals: real homeowner request
POSITIVE_SIGNALS = [
    # "Looking for/need/recommend" — asking for help
    (r"(?i)(looking\s+for|need\s+a|need\s+someone|recommend\s+a|recommendation)", +25),
    # "Anyone know/can someone" — community question
    (r"(?i)(anyone\s+know|can\s+someone|does\s+anyone|has\s+anyone|who\s+do\s+you)", +25),
    # Problem language — something is broken (require context: "my X is broken", "X is leaking", etc.)
    (r"(?i)(my|our|the|is|are|has|have|got)\s+\w{0,20}\s*(broken|damaged|leaking|cracked|rotting|rotted|peeling|warped|sagging|clogged|flooded|mold|mildew)", +30),
    (r"(?i)(broken|damaged|leaking|cracked|flooded)\s+(pipe|faucet|toilet|window|door|wall|ceiling|floor|roof|furnace|boiler|water\s*heater|sink|shower|tub)", +30),
    # First-person perspective — their own home
    (r"(?i)(my\s+(kitchen|bathroom|basement|house|home|unit|apartment|condo|bedroom|living\s*room|garage|roof|deck|patio|floor|wall|ceiling|door|window|pipe|faucet|shower|tub|toilet|sink|drywall))", +20),
    (r"(?i)(our\s+(kitchen|bathroom|basement|house|home)|i\s+need\s+)", +20),
    # Genuine questions (not rhetorical ad questions like "Emergency? Remodeling?")
    (r"(?i)(anyone|how\s+much|what\s+should|where\s+can|who\s+(do|should|can)|is\s+(it|this|there)|should\s+I|can\s+(I|someone)|has\s+anyone|does\s+anyone|do\s+you|need\s+help|any\s+(advice|tips|recommend))\s*\??", +15),
    # Budget questions
    (r"(?i)(how\s+much|what\s+should\s+(i|we)\s+(pay|expect|budget)|cost\s+estimate|ballpark|quote\s+for)", +20),
    # Timeline language — planning a project
    (r"(?i)(this\s+week|next\s+week|this\s+month|soon|planning\s+to|about\s+to|getting\s+ready|want\s+to\s+(start|get|have|do))", +15),
    # Urgency — real emergency (require first-person or question context)
    (r"(?i)(I|we|my|our)\s+\w{0,15}\s*(asap|urgently|emergency|urgent)", +25),
    (r"(?i)(water\s+damage|pipe\s+burst|flood(ed|ing)?)\s+(in|at|from|my|our)", +25),
    # Ready to hire
    (r"(?i)(ready\s+to\s+(start|hire|begin)|want\s+to\s+hire)", +20),
    # Reddit/Facebook source bonus (already curated by scanner)
    # Applied in score_post() based on source_type, not regex
]

# Source bonuses (applied on top of text scoring)
SOURCE_BONUS = {
    "reddit": +15,
    "facebook": +15,
    "nextdoor": +10,
    "craigslist": 0,
}

# Nextdoor-specific noise (directory pages, not real posts)
NEXTDOOR_NOISE_URLS = [
    r"/pages/",
    r"/find/",
    r"/search/",
]


# ── Scoring ───────────────────────────────────────────────────────────

def score_post(post, source_type):
    """
    Score a post for homeowner intent. Returns (score, breakdown).
    Score range: -100 to +100. Only score >= SCORE_THRESHOLD gets logged.
    """
    title = post.get("title", "")
    snippet = post.get("snippet", "")
    url = post.get("url", "")
    combined = f"{title} {snippet}"

    score = 0
    breakdown = []

    # Nextdoor directory pages are instant-skip
    if source_type == "nextdoor":
        for pattern in NEXTDOOR_NOISE_URLS:
            if re.search(pattern, url):
                return -100, ["Nextdoor directory page"]
        if re.match(r"^[A-Z].*- (Chicago|.+), IL - Nextdoor$", title):
            return -100, ["Nextdoor business listing"]

    # Check negative signals
    for pattern, penalty in NEGATIVE_SIGNALS:
        if re.search(pattern, combined):
            score += penalty
            breakdown.append(f"{penalty:+d} {pattern[:40]}")

    # Check positive signals
    for pattern, bonus in POSITIVE_SIGNALS:
        if re.search(pattern, combined):
            score += bonus
            breakdown.append(f"{bonus:+d} {pattern[:40]}")

    # Source bonus
    bonus = SOURCE_BONUS.get(source_type, 0)
    if bonus:
        score += bonus
        breakdown.append(f"{bonus:+d} source:{source_type}")

    # Single project focus bonus: only one category detected = focused request
    # Only award this if there's already a positive signal (prevents contractor ads
    # that happen to mention one service from crossing the threshold)
    has_positive = any(re.search(p, combined) for p, b in POSITIVE_SIGNALS if b > 0)
    if has_positive:
        categories_found = sum(1 for p, _ in CATEGORY_RULES if re.search(p, combined))
        if categories_found == 1:
            score += 10
            breakdown.append("+10 single project focus")

    # Clamp to range
    score = max(-100, min(100, score))

    return score, breakdown


def score_to_priority(score, source_type="reddit"):
    """Map homeowner score to priority label. Uses source-specific threshold."""
    threshold = SCORE_THRESHOLD_BY_SOURCE.get(source_type, SCORE_THRESHOLD)
    
    if score >= 50:
        return "\U0001f525 Hot"   # fire emoji
    if score >= 25:
        return "\u26a1 Warm"     # lightning emoji
    if score >= threshold:
        return "\u2744\ufe0f Cold"  # snowflake emoji
    return None  # Below threshold — don't log


# ── Category Detection ─────────────────────────────────────────────────

CATEGORY_RULES = [
    (r"(?i)kitchen.*(remodel|renovat|gut|cabinet|countertop)", "Kitchen Remodel"),
    (r"(?i)bath(room)?.*(remodel|renovat|gut|tile|shower|tub)", "Bathroom Remodel"),
    (r"(?i)(full|whole|complete).*(remodel|renovat|gut)", "Full Renovation"),
    (r"(?i)(floor|lvp|laminate|hardwood|vinyl|carpet)", "Flooring"),
    (r"(?i)(paint|painter|painting|repaint)", "Painting"),
    (r"(?i)(drywall|sheetrock|plaster|skim.?coat)", "Drywall"),
    (r"(?i)(tile|tiling|backsplash|grout)", "Tile"),
    (r"(?i)(plumb|faucet|toilet|pipe|drain|water.?heater|sump)", "Plumbing"),
    (r"(?i)(electric|outlet|wiring|panel|switch|light.?fixture)", "Electrical"),
    (r"(?i)(basement|finish.*basement)", "Basement"),
    (r"(?i)(deck|patio|porch|fence|fencing)", "Deck/Patio"),
    (r"(?i)(unit.?turn|tenant|rental|landlord|property.?manag)", "Unit Turnover"),
    (r"(?i)(commercial|office|retail|store.?front)", "Commercial"),
    (r"(?i)(handyman|handy.?man|odd.?job|small.?job|fix|repair|maintenance)", "Handyman"),
    (r"(?i)(remodel|renovat|home.?improv|construction|contractor|general.?contract)", "Full Renovation"),
    (r"(?i)(roof|gutter|siding|window|door|insulation)", "Other"),
]


def detect_category(text):
    """Auto-detect project category from title + snippet."""
    for pattern, category in CATEGORY_RULES:
        if re.search(pattern, text):
            return category
    return "Other"


# ── Tracker Helpers ────────────────────────────────────────────────────

def get_existing_urls(ws):
    """Get all URLs already in the tracker (stored in notes column P)."""
    urls = set()
    for row in range(3, ws.max_row + 1):
        notes = ws[f"P{row}"].value
        desc = ws[f"I{row}"].value
        if notes:
            urls.add(str(notes).strip())
        if desc:
            urls.add(str(desc).strip())
    return urls


def find_next_row(ws):
    """Find the next empty row in the tracker."""
    for row in range(3, 1003):
        if ws[f"B{row}"].value is None:
            return row
    return ws.max_row + 1


def normalize_source(post_source, source_type):
    """Clean up source name for display in dashboard."""
    if source_type == "reddit":
        return post_source  # Already "Reddit - r/subreddit"
    if source_type == "nextdoor":
        return "Nextdoor"
    if source_type == "craigslist":
        return "Craigslist"
    if source_type == "facebook":
        return "Facebook"
    return post_source or source_type.capitalize()


# ── Processing ─────────────────────────────────────────────────────────

def process_source(source_type, filepath, ws, existing_urls, dry_run=False):
    """Process a single scanner JSON file. Score each post and write qualified leads."""
    if not os.path.exists(filepath):
        return 0

    with open(filepath, "r", encoding="utf-8-sig") as f:
        data = json.load(f)

    # Facebook outputs a flat array, others use {posts: [...]}
    if source_type == "facebook":
        posts = data if isinstance(data, list) else data.get("posts", [])
    else:
        posts = data.get("posts", [])

    if not posts:
        return 0

    logged = 0
    skipped = 0
    next_row = find_next_row(ws) if not dry_run else 0

    for post in posts:
        url = post.get("url", "")
        title = post.get("title", "")
        snippet = post.get("snippet", "")

        # Facebook sometimes has no title
        if source_type == "facebook" and not title:
            title = snippet

        # Skip if already in tracker
        if url and url in existing_urls:
            continue

        # Score the post
        score, breakdown = score_post(post, source_type)
        priority = score_to_priority(score, source_type)

        if dry_run:
            status = "PASS" if priority else "SKIP"
            threshold = SCORE_THRESHOLD_BY_SOURCE.get(source_type, SCORE_THRESHOLD)
            print(f"  [{status}] score={score:+4d} (threshold={threshold:+d}) | {title[:70]}")
            if priority:
                logged += 1
            else:
                skipped += 1
            continue

        # Below threshold — skip
        if not priority:
            skipped += 1
            continue

        # Build description
        desc = title
        if snippet and snippet != title:
            desc = f"{title} — {snippet}"
        if len(desc) > 300:
            desc = desc[:297] + "..."

        # Auto-detect fields
        combined = f"{title} {snippet}"
        category = detect_category(combined)
        source = normalize_source(post.get("source", ""), source_type)
        location = "Chicago"

        # Extract location from Nextdoor if present
        if source_type == "nextdoor":
            loc_match = re.search(r"- ([A-Z][a-z]+(?:\s[A-Z][a-z]+)*), IL", title)
            if loc_match:
                location = loc_match.group(1) + ", IL"

        # Write to tracker
        ws[f"B{next_row}"] = datetime.now()
        ws[f"B{next_row}"].number_format = "MM/DD/YYYY"
        ws[f"C{next_row}"] = source
        ws[f"D{next_row}"] = category
        ws[f"E{next_row}"] = ""  # Client name unknown
        ws[f"F{next_row}"] = ""  # Phone unknown
        ws[f"G{next_row}"] = ""  # Email unknown
        ws[f"H{next_row}"] = location
        ws[f"I{next_row}"] = desc
        ws[f"J{next_row}"] = ""  # Est value - Hunter will fill
        ws[f"K{next_row}"] = priority
        ws[f"L{next_row}"] = "New"
        ws[f"O{next_row}"] = "No"
        ws[f"P{next_row}"] = url

        # Queue for Karl if score >= 25 (Warm or Hot)
        lead_name = post.get("author", "") or post.get("name", "") or ""
        if queue_for_karl(url, source_type, lead_name, score):
            pass  # Silently queued

        existing_urls.add(url)
        next_row += 1
        logged += 1

    if dry_run:
        print(f"  -> {logged} would pass, {skipped} would skip")
    return logged


# ── Agent Queue ───────────────────────────────────────────────────────

def queue_for_karl(url, source_type, lead_name, score):
    """
    Write qualified leads (score >= 25) to agent-queue.json for Karl to draft replies.
    Deduplicates by postUrl to avoid double-queuing.
    """
    if score < 25:
        return False

    # Read existing queue
    if os.path.exists(AGENT_QUEUE_PATH):
        with open(AGENT_QUEUE_PATH, "r", encoding="utf-8") as f:
            try:
                queue = json.load(f)
            except json.JSONDecodeError:
                queue = []
    else:
        queue = []

    # Deduplicate by postUrl
    existing_urls = {item.get("postUrl") for item in queue}
    if url in existing_urls:
        return False

    # Determine source label
    source_map = {
        "facebook": "facebook",
        "reddit": "reddit",
        "nextdoor": "nextdoor",
        "craigslist": "craigslist",
    }

    now = datetime.utcnow().isoformat() + "Z"
    queue_item = {
        "id": str(uuid.uuid4()),
        "source": source_map.get(source_type, source_type),
        "leadName": lead_name or "(unnamed)",
        "postUrl": url,
        "score": score,
        "status": "new",
        "draftReply": None,
        "assignedTo": "karl",
        "createdAt": now,
        "updatedAt": now,
    }

    queue.append(queue_item)

    with open(AGENT_QUEUE_PATH, "w", encoding="utf-8") as f:
        json.dump(queue, f, indent=2)

    return True


def main():
    parser = argparse.ArgumentParser(description="Integrate scanner leads into Mission Control")
    parser.add_argument("--source", choices=list(SCANNER_FILES.keys()),
                        help="Process only this source (default: all)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Score all posts and show results without writing to tracker")
    args = parser.parse_args()

    if not args.dry_run:
        if not os.path.exists(TRACKER_PATH):
            print(f"ERROR: Tracker not found at {TRACKER_PATH}")
            sys.exit(1)

        wb = load_workbook(TRACKER_PATH)
        ws = wb["Lead Tracker"]
        existing_urls = get_existing_urls(ws)
    else:
        wb = ws = None
        existing_urls = set()
        print("=== DRY RUN — scoring only, no writes ===\n")

    sources = {args.source: SCANNER_FILES[args.source]} if args.source else SCANNER_FILES
    total = 0

    for source_type, filepath in sources.items():
        if args.dry_run:
            print(f"\n--- {source_type.upper()} ---")
        count = process_source(source_type, filepath, ws, existing_urls, dry_run=args.dry_run)
        if count > 0 and not args.dry_run:
            print(f"  {source_type}: {count} new leads logged")
        total += count

    if args.dry_run:
        print(f"\n=== Total: {total} leads would be logged ===")
        return

    if total > 0:
        wb.save(TRACKER_PATH)
        print(f"\nTotal: {total} new leads added to Mission Control")

        # Auto-sync office dashboard
        import subprocess
        sync_script = r"C:\Users\trisd\clawd\scripts\sync-office-data.py"
        if os.path.exists(sync_script):
            subprocess.run([
                r"C:\Users\trisd\AppData\Local\Programs\Python\Python312\python.exe",
                sync_script
            ], capture_output=True)
    else:
        print("No new leads to add.")


if __name__ == "__main__":
    main()
