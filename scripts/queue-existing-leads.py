#!/usr/bin/env python3
"""
Queue existing leads from tracker to Karl for reply drafting.
Use when leads are already logged but not yet queued for Karl.
"""

import json
import uuid
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(r"C:\Users\trisd\clawd\data")
TRACKER_PATH = DATA_DIR / "Valencia-Lead-Tracker.xlsx"
AGENT_QUEUE = DATA_DIR / "agent-queue.json"

def queue_tracker_leads():
    """Read New leads from tracker and queue for Karl."""
    
    if not TRACKER_PATH.exists():
        print("ERROR: Tracker not found")
        return 0
    
    # Load existing queue
    if AGENT_QUEUE.exists():
        with open(AGENT_QUEUE) as f:
            queue = json.load(f)
    else:
        queue = []
    
    existing_urls = {item.get("postUrl") for item in queue}
    queued_count = 0
    
    # Read tracker
    wb = load_workbook(TRACKER_PATH)
    ws = wb["Lead Tracker"]
    
    for row in range(3, 505):
        status = ws[f"L{row}"].value
        
        # Only queue "New" leads (not yet contacted)
        if status != "New":
            continue
        
        # Get lead data from tracker
        date_cell = ws[f"B{row}"].value
        source = ws[f"C{row}"].value or "Unknown"
        desc = ws[f"I{row}"].value or ""
        url = ws[f"P{row}"].value or ""
        
        if not url or url in existing_urls:
            continue
        
        # Determine score based on source/desc (simplified)
        # Hot = Facebook, Reddit, specific project mention
        # Warm = Craigslist with multiple indicators
        # Cold = generic
        
        score = 20  # Default Warm
        
        if "facebook" in source.lower():
            score = 50  # Hot
        elif "reddit" in source.lower():
            score = 45  # Hot
        elif "nextdoor" in source.lower():
            score = 35  # Warm
        
        # Boost for specific project types
        if any(word in desc.lower() for word in ["kitchen", "bathroom", "remodel", "renovation"]):
            score += 10
        
        # Create queue item
        now = datetime.utcnow().isoformat() + "Z"
        queue_item = {
            "id": str(uuid.uuid4()),
            "source": source,
            "leadName": "(from tracker)",
            "postUrl": url,
            "score": min(100, score),
            "status": "new",
            "draftReply": None,
            "assignedTo": "karl",
            "createdAt": now,
            "updatedAt": now,
        }
        
        queue.append(queue_item)
        existing_urls.add(url)
        queued_count += 1
        
        print(f"  Queued: {desc[:60]}")
    
    # Save updated queue
    with open(AGENT_QUEUE, "w") as f:
        json.dump(queue, f, indent=2)
    
    return queued_count

if __name__ == "__main__":
    print("\n" + "="*70)
    print("QUEUEING EXISTING TRACKER LEADS FOR KARL")
    print("="*70)
    
    count = queue_tracker_leads()
    
    print(f"\n[OK] Queued {count} leads for Karl")
    print(f"  Karl will now draft replies and send approval requests")
    print("\n" + "="*70 + "\n")
