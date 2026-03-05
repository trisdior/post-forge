#!/usr/bin/env python3
"""
Final Lead Hunter - Working Craigslist scraper with strict validation
"""

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import URLError

# Configuration
SEEN_FILE = Path(r'C:\Users\trisd\clawd\data\craigslist-seen.json')
REPORT_FILE = Path(r'C:\Users\trisd\clawd\data\lead-hunt-report.json')
TRACKER_PATH = Path(r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx')

# Red flags - CONTRACTOR LANGUAGE (Strict homeowner filtering)
RED_FLAGS_STRICT = [
    # Contractor self-descriptions
    r'\b(we|us|our)\s+(offer|provide|do|specialize|handle|install|repair|service)\b',
    r'\b(our|our company|our team)\s+(can|will|is|are)',
    r'\blicensed\s+(and|&)?\s+insured\b',
    r'\bcall\s+(us|me|for)\b',
    r'\bcontact\s+(us|me|for)\b',
    r'\byears?\s+of\s+experience\b',
    r'\bfamily\s+owned\b',
    r'\b(professional|certified|experienced|licensed)\s+(contractor|builder)',
    # Company/business indicators
    r'\b(llc|inc|corp|co\.?|company|services|construction|group|enterprises?)\s*$',
    r'\b(free estimates?|free quotes?|free consultations?)\b',
    # Employment/job posting language
    r'\b(hiring|employment|apply|position|job opening|w2|1099)\b',
    # Selling/marketing language
    r'(visit|check out|follow|see|look at)\s+(us|our|website|page|facebook|instagram)',
    # Multiple service categories (contractor doing multiple things)
    r'(painting|plumbing|electrical|hvac|roofing|flooring|kitchen|bathroom)',  # Need 2+ of these
]

# Homeowner POSITIVE signals
HOMEOWNER_SIGNALS = [
    (r'\b(i\s+)?need\b', 'expressing need'),
    (r'\b(my|our|the)\s+(home|house|apartment|condo|place)\b', 'describes own property'),
    (r'\b(damage|broken|leaking|cracked|water\s+damage|mold|stain)\b', 'damage described'),
    (r'\b(urgent|asap|immediately|soon|before\s+(next|this)\s+(week|month))\b', 'time-sensitive'),
    (r'\b(budget|quote|price|cost|afford|estimate)\b', 'budget mentioned'),
    (r'\b(looking\s+for|seeking|need|can\s+you|do\s+you|would\s+you)\b', 'actively asking'),
    (r'\b(small|simple|quick)\s+(job|repair|fix)\b', 'minor work'),
    (r'\b(in\s+)?chicago|chicago area\b', 'location specific'),
]

def load_seen():
    """Load already-seen post IDs"""
    if SEEN_FILE.exists():
        try:
            with open(SEEN_FILE) as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_seen(seen):
    """Save seen post IDs"""
    SEEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SEEN_FILE, 'w') as f:
        json.dump(seen, f, indent=2)

def fetch_page(url, timeout=30):
    """Fetch webpage content"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        req = Request(url, headers=headers)
        with urlopen(req, timeout=timeout) as response:
            return response.read().decode('utf-8', errors='ignore')
    except Exception as e:
        return None

def extract_posts(html):
    """Extract individual posts from listing page"""
    posts = []
    
    # Parse cl-static-search-result entries
    pattern = r'<li[^>]*class="[^"]*cl-static-search-result[^"]*"[^>]*>.*?<a\s+href="([^"]+)"[^>]*>\s*<div[^>]*class="[^"]*title[^"]*">([^<]+)</div>'
    
    matches = re.finditer(pattern, html, re.DOTALL | re.IGNORECASE)
    for m in matches:
        url, title = m.groups()
        if '//' not in url:
            url = f"https://chicago.craigslist.org{url}"
        
        # Extract post ID
        id_match = re.search(r'/(\d+)\.html', url)
        post_id = id_match.group(1) if id_match else None
        
        if post_id:
            posts.append({
                'postId': post_id,
                'postUrl': url,
                'title': title.strip()[:100]
            })
    
    return posts

def validate_post_strict(post_id, post_url, title):
    """Strict homeowner validation - high bar for contractor filtering"""
    html = fetch_page(post_url)
    if not html:
        return None
    
    # Extract post body
    body_match = re.search(
        r'<section id="postingbody"[^>]*>(.+?)</section>',
        html,
        re.DOTALL
    )
    
    full_text = body_match.group(1) if body_match else title
    
    # Clean HTML
    full_text = re.sub(r'<[^>]+>', ' ', full_text)
    full_text = re.sub(r'&[a-z]+;', '', full_text)
    full_text = re.sub(r'\s+', ' ', full_text).strip()
    
    # Check text
    check_text = f"{title} {full_text}".lower()
    
    result = {
        'postId': post_id,
        'postUrl': post_url,
        'title': title,
        'textSnippet': full_text[:250],
        'valid': True,
        'reasons': [],
        'signals': [],
        'redFlagsFound': []
    }
    
    # RED FLAG CHECK - be STRICT
    contractor_service_count = 0
    for flag_pattern in RED_FLAGS_STRICT:
        if re.search(flag_pattern, check_text, re.IGNORECASE):
            # Special handling: service words in list
            if flag_pattern == r'(painting|plumbing|electrical|hvac|roofing|flooring|kitchen|bathroom)':
                contractor_service_count += 1
            else:
                result['valid'] = False
                result['redFlagsFound'].append(flag_pattern[:35])
    
    # If multiple service types mentioned, likely a contractor
    if contractor_service_count >= 2:
        result['valid'] = False
        result['reasons'].append(f"Multiple service categories ({contractor_service_count}) - contractor language")
    
    # HOMEOWNER SIGNAL CHECK
    signal_count = 0
    for signal_pattern, signal_label in HOMEOWNER_SIGNALS:
        if re.search(signal_pattern, check_text, re.IGNORECASE):
            signal_count += 1
            result['signals'].append(signal_label)
    
    # Final decision
    if result['valid']:
        if signal_count == 0:
            result['valid'] = False
            result['reasons'].append("No homeowner signals")
        elif signal_count < 2:
            result['reasons'].append(f"Weak ({signal_count} signal) - skipping for safety")
            result['valid'] = False
        else:
            result['reasons'].append(f"Strong homeowner lead ({signal_count} signals)")
    
    return result

def log_lead_to_tracker(lead):
    """Log valid lead to Excel tracker"""
    try:
        from openpyxl import load_workbook
        
        if not TRACKER_PATH.exists():
            print(f"    [!] Tracker not found at {TRACKER_PATH}")
            return False
        
        wb = load_workbook(TRACKER_PATH)
        
        # Find the right sheet
        sheet_name = None
        for name in ['Facebook Leads', 'Lead Tracker']:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"    [!] No lead sheet found (available: {wb.sheetnames})")
            return False
        
        ws = wb[sheet_name]
        
        # Find next empty row
        next_row = 3
        for row in range(3, 1000):
            if ws[f'B{row}'].value is None:
                next_row = row
                break
        
        # Write lead
        ws[f'B{next_row}'] = datetime.now()
        ws[f'C{next_row}'] = 'Craigslist'
        ws[f'D{next_row}'] = 'Home Services'
        ws[f'I{next_row}'] = lead['title'][:100]
        ws[f'K{next_row}'] = 'Hot'
        ws[f'L{next_row}'] = 'New'
        ws[f'P{next_row}'] = lead['postUrl']
        
        wb.save(TRACKER_PATH)
        return True
    except Exception as e:
        print(f"    [!] Tracker error: {e}")
        return False

def main():
    # Fix Windows console encoding
    import sys
    import io
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    
    print("\n" + "="*70)
    print("STRICT LEAD HUNTER - HOMEOWNER VALIDATION")
    print("="*70)
    
    seen = load_seen()
    valid_leads = []
    rejected = []
    checked = 0
    
    # Search in Craigslist household services section
    search_url = "https://chicago.craigslist.org/search/hss"
    
    print(f"\n[*] Fetching listings from: {search_url}")
    html = fetch_page(search_url)
    
    if not html:
        print("[!] Failed to fetch Craigslist page")
        return
    
    # Extract posts
    posts = extract_posts(html)
    print(f"[+] Found {len(posts)} listings on page\n")
    
    # Validate each post
    for i, post in enumerate(posts, 1):
        if post['postId'] in seen:
            continue
        
        checked += 1
        print(f"[{i:3d}] Validating: {post['title'][:50]}...")
        
        validation = validate_post_strict(
            post['postId'],
            post['postUrl'],
            post['title']
        )
        
        if validation:
            if validation['valid']:
                print(f"      [ACCEPT] {validation['reasons'][0]}")
                print(f"      Signals: {', '.join(validation['signals'])}")
                valid_leads.append(validation)
                seen[post['postId']] = datetime.now().isoformat()
                
                # Log to tracker
                if log_lead_to_tracker(validation):
                    print(f"      [+] Added to tracker")
            else:
                # Print first rejection reason
                reason = validation['reasons'][0] if validation['reasons'] else "Rejected"
                print(f"      [SKIP] {reason}")
                rejected.append(validation)
        
        time.sleep(0.2)
    
    # Save tracking
    save_seen(seen)
    
    # Generate report
    quality = (len(valid_leads) / checked * 100) if checked > 0 else 0
    
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'totalChecked': checked,
            'validLeadsFound': len(valid_leads),
            'rejectedLeads': len(rejected),
            'qualityScorePercent': round(quality, 1)
        },
        'validLeads': valid_leads,
        'rejectedLeads': rejected[:10]
    }
    
    # Save report
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, indent=2, default=str)
    
    # Print summary
    print("\n" + "="*70)
    print(f"HUNT COMPLETE - Quality Score: {quality}%")
    print("="*70)
    print(f"Posts Checked:    {checked}")
    print(f"Valid Leads:      {len(valid_leads)}")
    print(f"Rejected:         {len(rejected)}")
    print(f"Report saved:     {REPORT_FILE}")
    
    if valid_leads:
        print(f"\n{len(valid_leads)} HOMEOWNER LEADS FOUND:")
        for j, lead in enumerate(valid_leads, 1):
            print(f"\n  {j}. {lead['title']}")
            print(f"     URL: {lead['postUrl']}")
            print(f"     Signals: {len(lead['signals'])} - {', '.join(lead['signals'][:3])}")
    
    print("\n" + "="*70 + "\n")

if __name__ == '__main__':
    main()
