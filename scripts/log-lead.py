"""
Log a lead to the Valencia Lead Tracker spreadsheet.
Usage: python log-lead.py "source" "category" "title/description" "location" "url" "est_value"
"""
import sys
import os
from datetime import datetime
from openpyxl import load_workbook

TRACKER_PATH = r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx'

def log_lead(source, category, description, location, url, est_value=""):
    wb = load_workbook(TRACKER_PATH)
    # Dashboard reads from "Lead Tracker" sheet
    ws = wb["Lead Tracker"]
    
    # Find next empty row
    next_row = 3
    for row in range(3, 503):
        if ws[f'B{row}'].value is None:
            next_row = row
            break
    
    # Write lead data
    ws[f'B{next_row}'] = datetime.now()
    ws[f'B{next_row}'].number_format = 'MM/DD/YYYY'
    ws[f'C{next_row}'] = source
    ws[f'D{next_row}'] = category
    ws[f'E{next_row}'] = ""  # Client name - unknown from CL
    ws[f'F{next_row}'] = ""  # Phone - unknown
    ws[f'G{next_row}'] = ""  # Email - unknown
    ws[f'H{next_row}'] = location
    ws[f'I{next_row}'] = description
    
    if est_value:
        try:
            ws[f'J{next_row}'] = int(est_value.replace('$','').replace(',',''))
            ws[f'J{next_row}'].number_format = '$#,##0'
        except:
            pass
    
    ws[f'K{next_row}'] = '⚡ Warm'  # Default priority
    ws[f'L{next_row}'] = 'New'
    ws[f'O{next_row}'] = 'No'
    ws[f'P{next_row}'] = url  # Store URL in notes
    
    wb.save(TRACKER_PATH)
    print(f"Lead #{next_row - 2} logged: {description[:50]}")
    
    # Auto-sync office dashboard
    import subprocess
    subprocess.run([
        r'C:\Users\trisd\AppData\Local\Programs\Python\Python312\python.exe',
        r'C:\Users\trisd\clawd\scripts\sync-office-data.py'
    ], capture_output=True)

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Usage: python log-lead.py source category description location [url] [est_value]")
        sys.exit(1)
    
    source = sys.argv[1]
    category = sys.argv[2]
    description = sys.argv[3]
    location = sys.argv[4]
    url = sys.argv[5] if len(sys.argv) > 5 else ""
    est_value = sys.argv[6] if len(sys.argv) > 6 else ""
    
    log_lead(source, category, description, location, url, est_value)
