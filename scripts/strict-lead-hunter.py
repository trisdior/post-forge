#!/usr/bin/env python3
"""
Strict Lead Hunter for Valencia Construction
Validates leads against homeowner vs contractor criteria
Only adds REAL homeowner leads
"""

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import quote
from urllib.error import URLError, HTTPError

# Configuration
SEEN_FILE = Path(r'C:\Users\trisd\clawd\data\craigslist-seen.json')
REPORT_FILE = Path(r'C:\Users\trisd\clawd\data\lead-hunt-report.json')
TRACKER_PATH = Path(r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx')

# Search terms targeting homeowners asking for help
SEARCH_TERMS = [
    "need painter",
    "looking for painter",
    "help with kitchen",
    "kitchen work needed",
    "bathroom repairs needed",
    "bathroom work",
    "drywall repair",
    "wall damage needs fixing",
    "flooring install needed"
]

# Red flags - skip these
RED_FLAGS = [
    r'\bwe offer\b',
    r'\bwe specialize\b',
    r'\bwe do\b',
    r'\blicensed and insured\b',
    r'\byears of experience\b',
    r'\bfamily owned\b',
    r'\bcall us\b',
    r'\bcontact us\b',
    r'\w+\s+(?:llc|inc|co\.?|corp|construction|company)\b',
    r'\b(?:W2|1099)\b',
    r'\b(?:hiring|employment|job posting|position)\b'
]

# Homeowner indicators
HOMEOWNER_PATTERNS = [
    (r'\b(?:I|we|my|our|our home|our house|my house|my place)\b', 'personal pronouns'),
    (r'\b(?:need help|need someone|looking for|seeking|urgent)\b', 'requesting help'),
    (r'\b(?:damage|repair|fix|broken|leak|cracked|water damage)\b', 'specific issue'),
]

def load_seen():
    """Load already-seen post IDs"""
    if SEEN_FILE.exists():
        try:
            with open(SEEN_FILE) as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_seen(seen):
    """Save seen post IDs"""
    SEEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SEEN_FILE, 'w') as f:
        json.dump(seen, f, indent=2)

def search_craigslist(term):
    """Search Craigslist for a term"""
    encoded = quote(term)
    url = f"https://chicago.craigslist.org/search/hsh?query={encoded}"
    
    print(f"[*] Searching: '{term}'")
    results = []
    
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        req = Request(url, headers=headers)
        with urlopen(req, timeout=30) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Extract listing links
        pattern = r'<a href="(https://chicago\.craigslist\.org/[^"]*?/(\d+)\.html)"[^>]*>([^<]+)</a>'
        matches = re.finditer(pattern, html)
        
        for m in matches:
            post_url, post_id, title = m.groups()
            results.append({
                'postUrl': post_url,
                'postId': post_id,
                'title': title.strip()[:100]
            })
        
        print(f"    Found {len(results)} results")
    except Exception as e:
        print(f"    Error: {e}")
    
    return results

def validate_post(post_url, post_id, title):
    """Fetch and validate a post"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        req = Request(post_url, headers=headers)
        with urlopen(req, timeout=30) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # Extract post body
        body_match = re.search(
            r'<section id="postingbody"[^>]*>(.+?)</section>',
            html,
            re.DOTALL
        )
        
        if body_match:
            post_text = body_match.group(1)
        else:
            post_text = title
        
        # Clean HTML
        post_text = re.sub(r'<[^>]+>', ' ', post_text)
        post_text = re.sub(r'\s+', ' ', post_text)
        post_text = re.sub(r'&quot;', '"', post_text)
        post_text = re.sub(r'&amp;', '&', post_text)
        post_text = post_text.strip()
        
        # Validation result
        validation = {
            'isValid': True,
            'score': 0,
            'reason': '',
            'indicators': [],
            'redFlags': []
        }
        
        # Check RED FLAGS
        for pattern in RED_FLAGS:
            if re.search(pattern, post_text, re.IGNORECASE):
                validation['isValid'] = False
                validation['redFlags'].append(pattern[:30])
        
        # Check HOMEOWNER INDICATORS
        indicator_count = 0
        for pattern, label in HOMEOWNER_PATTERNS:
            if re.search(pattern, post_text, re.IGNORECASE):
                validation['indicators'].append(label)
                indicator_count += 1
                validation['score'] += 30
        
        # Final decision
        if validation['isValid']:
            if indicator_count < 2:
                validation['isValid'] = False
                validation['reason'] = f"Not enough homeowner signals ({indicator_count} found)"
            else:
                validation['reason'] = "Valid homeowner lead"
        else:
            validation['reason'] = f"Red flags: {', '.join(validation['redFlags'][:2])}"
        
        return {
            'postId': post_id,
            'title': title,
            'url': post_url,
            'textSnippet': post_text[:300],
            'validation': validation,
            'valid': validation['isValid']
        }
    
    except Exception as e:
        print(f"    Error fetching {post_id}: {e}")
        return None

def log_lead_to_tracker(lead):
    """Log a valid lead to the tracker spreadsheet"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(TRACKER_PATH)
        ws = wb["Facebook Leads"]
        
        # Find next empty row
        next_row = 3
        for row in range(3, 503):
            if ws[f'B{row}'].value is None:
                next_row = row
                break
        
        # Write lead data
        ws[f'B{next_row}'] = datetime.now()
        ws[f'B{next_row}'].number_format = 'MM/DD/YYYY'
        ws[f'C{next_row}'] = 'Craigslist'
        ws[f'D{next_row}'] = 'Home Remodel'
        ws[f'E{next_row}'] = ''  # Client name unknown
        ws[f'F{next_row}'] = ''  # Phone unknown
        ws[f'G{next_row}'] = ''  # Email unknown
        ws[f'H{next_row}'] = 'Chicago'
        ws[f'I{next_row}'] = lead['title']
        ws[f'K{next_row}'] = 'Hot'  # Default priority
        ws[f'L{next_row}'] = 'New'
        ws[f'O{next_row}'] = 'No'
        ws[f'P{next_row}'] = lead['url']
        
        wb.save(TRACKER_PATH)
        print(f"    Logged to tracker: {lead['title'][:40]}")
        return True
    except Exception as e:
        print(f"    Error logging to tracker: {e}")
        return False

def main():
    seen = load_seen()
    valid_leads = []
    rejected_leads = []
    total_checked = 0
    
    # Search with each term
    for term in SEARCH_TERMS:
        results = search_craigslist(term)
        time.sleep(1)
        
        for result in results:
            post_id = result['postId']
            
            # Skip if already seen
            if post_id in seen:
                continue
            
            total_checked += 1
            validation = validate_post(
                result['postUrl'],
                post_id,
                result['title']
            )
            
            if validation:
                if validation['valid']:
                    print(f"  [+] VALID: {result['title'][:60]}")
                    valid_leads.append(validation)
                    seen[post_id] = datetime.now().isoformat()
                    
                    # Log to tracker
                    log_lead_to_tracker(validation)
                else:
                    print(f"  [-] REJECTED: {result['title'][:60]}")
                    print(f"      Reason: {validation['validation']['reason']}")
                    rejected_leads.append(validation)
            
            time.sleep(0.5)
        
        time.sleep(1)
    
    # Save seen posts
    save_seen(seen)
    
    # Generate report
    quality_score = (len(valid_leads) / total_checked * 100) if total_checked > 0 else 0
    
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'totalSearched': total_checked,
            'validLeadsFound': len(valid_leads),
            'rejectedLeads': len(rejected_leads),
            'qualityScore': round(quality_score, 1)
        },
        'validLeads': valid_leads,
        'rejectedLeads': rejected_leads[:10]  # Top 10 for audit
    }
    
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, indent=2)
    
    # Print summary
    print("\n" + "="*50)
    print("LEAD HUNT SUMMARY")
    print("="*50)
    print(f"Total Posts Checked: {total_checked}")
    print(f"Valid Homeowner Leads Found: {len(valid_leads)}")
    print(f"Rejected Posts: {len(rejected_leads)}")
    print(f"Quality Score: {quality_score}%")
    
    if valid_leads:
        print("\nVALID LEADS:")
        for lead in valid_leads:
            print(f"  - {lead['title']}")
            print(f"    URL: {lead['url']}")
    
    print(f"\nReport saved to: {REPORT_FILE}")
    
    return json.dumps(report, indent=2)

if __name__ == '__main__':
    result = main()
    print("\n" + result)
