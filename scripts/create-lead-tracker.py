from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

wb = Workbook()

# ============ STYLES ============
BLACK_FILL = PatternFill('solid', fgColor='111111')
GOLD_FILL = PatternFill('solid', fgColor='D4A017')
DARK_GOLD_FILL = PatternFill('solid', fgColor='B8860B')
LIGHT_GRAY_FILL = PatternFill('solid', fgColor='F5F5F5')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
GREEN_FILL = PatternFill('solid', fgColor='D4EDDA')
YELLOW_FILL = PatternFill('solid', fgColor='FFF3CD')
RED_FILL = PatternFill('solid', fgColor='F8D7DA')

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
GOLD_HEADER_FONT = Font(name='Arial', size=11, bold=True, color='111111')
TITLE_FONT = Font(name='Arial', size=16, bold=True, color='D4A017')
BODY_FONT = Font(name='Arial', size=11, color='333333')
BOLD_FONT = Font(name='Arial', size=11, bold=True, color='333333')

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
CURRENCY_FMT = '$#,##0'
DATE_FMT = 'MM/DD/YYYY'
PCT_FMT = '0%'

# ============ SHEET 1: LEAD TRACKER ============
ws = wb.active
ws.title = "Lead Tracker"
ws.sheet_properties.tabColor = "D4A017"

# Title row
ws.merge_cells('A1:P1')
ws['A1'] = 'VALENCIA CONSTRUCTION — LEAD TRACKER'
ws['A1'].font = TITLE_FONT
ws['A1'].fill = BLACK_FILL
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 40

# Headers
headers = [
    ('A', 'Lead #', 8),
    ('B', 'Date Found', 14),
    ('C', 'Source', 14),
    ('D', 'Category', 16),
    ('E', 'Client Name', 18),
    ('F', 'Phone', 16),
    ('G', 'Email', 22),
    ('H', 'Location', 16),
    ('I', 'Project Description', 30),
    ('J', 'Est. Value', 13),
    ('K', 'Priority', 11),
    ('L', 'Status', 14),
    ('M', 'Date Contacted', 15),
    ('N', 'Follow-Up Date', 15),
    ('O', 'Quote Sent', 13),
    ('P', 'Notes', 30),
]

for col_letter, header, width in headers:
    cell = ws[f'{col_letter}2']
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = BLACK_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[2].height = 30

# Data validation dropdowns
# Source
source_dv = DataValidation(type="list", formula1='"Craigslist,Facebook,Nextdoor,Google,Website Form,Referral,Cold Call,Thumbtack,Angi,Yelp,Walk-in,Other"')
source_dv.error = "Pick a source"
source_dv.errorTitle = "Invalid Source"
ws.add_data_validation(source_dv)
source_dv.add(f'C3:C502')

# Category
cat_dv = DataValidation(type="list", formula1='"Kitchen Remodel,Bathroom Remodel,Full Renovation,Flooring,Painting,Drywall,Tile,Plumbing,Electrical,Basement,Deck/Patio,Handyman,Unit Turnover,Commercial,Other"')
ws.add_data_validation(cat_dv)
cat_dv.add(f'D3:D502')

# Priority
priority_dv = DataValidation(type="list", formula1='"🔥 Hot,⚡ Warm,❄️ Cold"')
ws.add_data_validation(priority_dv)
priority_dv.add(f'K3:K502')

# Status
status_dv = DataValidation(type="list", formula1='"New,Contacted,Consultation Scheduled,Quote Sent,Follow-Up,Negotiating,Won,Lost,No Response,Not Qualified"')
ws.add_data_validation(status_dv)
status_dv.add(f'L3:L502')

# Quote Sent
quote_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
ws.add_data_validation(quote_dv)
quote_dv.add(f'O3:O502')

# Format data rows
for row in range(3, 53):
    for col_letter, _, _ in headers:
        cell = ws[f'{col_letter}{row}']
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT if col_letter in ['E','G','H','I','P'] else CENTER
    
    # Lead # formula
    ws[f'A{row}'] = f'=IF(B{row}="","",ROW()-2)'
    ws[f'A{row}'].alignment = CENTER
    
    # Date formats
    ws[f'B{row}'].number_format = DATE_FMT
    ws[f'M{row}'].number_format = DATE_FMT
    ws[f'N{row}'].number_format = DATE_FMT
    
    # Currency
    ws[f'J{row}'].number_format = CURRENCY_FMT
    
    # Alternating row colors
    if row % 2 == 0:
        for col_letter, _, _ in headers:
            ws[f'{col_letter}{row}'].fill = LIGHT_GRAY_FILL

# Freeze header
ws.freeze_panes = 'A3'

# ============ SHEET 2: DASHBOARD ============
dash = wb.create_sheet("Dashboard")
dash.sheet_properties.tabColor = "D4A017"

dash.merge_cells('A1:F1')
dash['A1'] = 'VALENCIA CONSTRUCTION — LEAD DASHBOARD'
dash['A1'].font = TITLE_FONT
dash['A1'].fill = BLACK_FILL
dash['A1'].alignment = CENTER
dash.row_dimensions[1].height = 40

# KPI Cards
kpis = [
    ('A3', 'Total Leads', '=COUNTA(\'Lead Tracker\'!B3:B502)'),
    ('B3', 'New / Uncontacted', '=COUNTIF(\'Lead Tracker\'!L3:L502,"New")'),
    ('C3', 'Consultations Set', '=COUNTIF(\'Lead Tracker\'!L3:L502,"Consultation Scheduled")'),
    ('D3', 'Quotes Sent', '=COUNTIF(\'Lead Tracker\'!O3:O502,"Yes")'),
    ('E3', 'Won', '=COUNTIF(\'Lead Tracker\'!L3:L502,"Won")'),
    ('F3', 'Lost', '=COUNTIF(\'Lead Tracker\'!L3:L502,"Lost")'),
]

for cell_ref, label, formula in kpis:
    # Label
    label_cell = dash[cell_ref]
    label_cell.value = label
    label_cell.font = GOLD_HEADER_FONT
    label_cell.fill = GOLD_FILL
    label_cell.alignment = CENTER
    label_cell.border = THIN_BORDER
    dash.column_dimensions[cell_ref[0]].width = 20
    
    # Value
    row_num = int(cell_ref[1]) + 1
    val_cell = dash[f'{cell_ref[0]}{row_num}']
    val_cell.value = formula
    val_cell.font = Font(name='Arial', size=24, bold=True, color='111111')
    val_cell.alignment = CENTER
    val_cell.border = THIN_BORDER
    dash.row_dimensions[row_num].height = 50

# Pipeline Value
dash['A6'] = 'Pipeline Value (All Active)'
dash['A6'].font = BOLD_FONT
dash.merge_cells('A6:C6')
dash['A7'] = '=SUMPRODUCT((\'Lead Tracker\'!L3:L502<>"Won")*(\'Lead Tracker\'!L3:L502<>"Lost")*(\'Lead Tracker\'!L3:L502<>"No Response")*(\'Lead Tracker\'!L3:L502<>"Not Qualified")*(\'Lead Tracker\'!J3:J502))'
dash['A7'].font = Font(name='Arial', size=28, bold=True, color='D4A017')
dash['A7'].number_format = '$#,##0'
dash.merge_cells('A7:C7')

dash['D6'] = 'Won Revenue'
dash['D6'].font = BOLD_FONT
dash.merge_cells('D6:F6')
dash['D7'] = '=SUMPRODUCT((\'Lead Tracker\'!L3:L502="Won")*(\'Lead Tracker\'!J3:J502))'
dash['D7'].font = Font(name='Arial', size=28, bold=True, color='228B22')
dash['D7'].number_format = '$#,##0'
dash.merge_cells('D7:F7')

# Conversion rate
dash['A9'] = 'Close Rate'
dash['A9'].font = BOLD_FONT
dash['A10'] = '=IF(A4=0,0,E4/A4)'
dash['A10'].font = Font(name='Arial', size=20, bold=True, color='111111')
dash['A10'].number_format = '0.0%'

dash['B9'] = 'Avg Deal Size'
dash['B9'].font = BOLD_FONT
dash['B10'] = '=IF(E4=0,0,D7/E4)'
dash['B10'].font = Font(name='Arial', size=20, bold=True, color='111111')
dash['B10'].number_format = '$#,##0'

# Leads by Source
dash['A12'] = 'LEADS BY SOURCE'
dash['A12'].font = HEADER_FONT
dash['A12'].fill = BLACK_FILL
dash.merge_cells('A12:C12')

sources = ['Craigslist','Facebook','Nextdoor','Google','Website Form','Referral','Cold Call','Thumbtack','Other']
for i, src in enumerate(sources):
    row = 13 + i
    dash[f'A{row}'] = src
    dash[f'A{row}'].font = BODY_FONT
    dash[f'A{row}'].border = THIN_BORDER
    dash[f'B{row}'] = f'=COUNTIF(\'Lead Tracker\'!C3:C502,"{src}")'
    dash[f'B{row}'].font = BOLD_FONT
    dash[f'B{row}'].alignment = CENTER
    dash[f'B{row}'].border = THIN_BORDER

# Leads by Category
dash['D12'] = 'LEADS BY CATEGORY'
dash['D12'].font = HEADER_FONT
dash['D12'].fill = BLACK_FILL
dash.merge_cells('D12:F12')

categories = ['Kitchen Remodel','Bathroom Remodel','Full Renovation','Flooring','Painting','Drywall','Tile','Plumbing','Electrical','Basement','Handyman','Unit Turnover','Other']
for i, cat in enumerate(categories):
    row = 13 + i
    dash[f'D{row}'] = cat
    dash[f'D{row}'].font = BODY_FONT
    dash[f'D{row}'].border = THIN_BORDER
    dash[f'E{row}'] = f'=COUNTIF(\'Lead Tracker\'!D3:D502,"{cat}")'
    dash[f'E{row}'].font = BOLD_FONT
    dash[f'E{row}'].alignment = CENTER
    dash[f'E{row}'].border = THIN_BORDER

dash.freeze_panes = 'A3'

# ============ SHEET 3: FOLLOW-UP QUEUE ============
fq = wb.create_sheet("Follow-Up Queue")
fq.sheet_properties.tabColor = "E85D04"

fq.merge_cells('A1:H1')
fq['A1'] = 'FOLLOW-UP QUEUE — Leads Needing Action'
fq['A1'].font = TITLE_FONT
fq['A1'].fill = BLACK_FILL
fq['A1'].alignment = CENTER
fq.row_dimensions[1].height = 40

fq_headers = [
    ('A', 'Lead #', 8),
    ('B', 'Client Name', 18),
    ('C', 'Category', 16),
    ('D', 'Status', 14),
    ('E', 'Follow-Up Date', 15),
    ('F', 'Phone', 16),
    ('G', 'Est. Value', 13),
    ('H', 'Notes', 35),
]

for col_letter, header, width in fq_headers:
    cell = fq[f'{col_letter}2']
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = DARK_GOLD_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    fq.column_dimensions[col_letter].width = width

# Instructions
fq['A3'] = 'Filter the Lead Tracker by Follow-Up Date to populate this view. Use Data > Sort & Filter on the main sheet.'
fq['A3'].font = Font(name='Arial', size=11, italic=True, color='888888')
fq.merge_cells('A3:H3')

fq.freeze_panes = 'A3'

# ============ SAVE ============
output_path = r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
