#!/usr/bin/env python3
"""
Flexible Lead Hunter - Broader search with strict validation
Searches Craigslist service categories and validates post content
"""

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

# Configuration
SEEN_FILE = Path(r'C:\Users\trisd\clawd\data\craigslist-seen.json')
REPORT_FILE = Path(r'C:\Users\trisd\clawd\data\lead-hunt-report.json')
TRACKER_PATH = Path(r'C:\Users\trisd\clawd\data\Valencia-Lead-Tracker.xlsx')

# Search terms - homeowners ASKING for help (not offering services)
SEARCH_PHRASES = [
    'painter needed',
    'kitchen repair',
    'bathroom repair',
    'drywall',
    'flooring',
    'home repair',
    'need contractor'
]

# Red flags - CONTRACTOR LANGUAGE
RED_FLAGS = [
    r'\b(we|us|our)\s+(offer|provide|do|specialize|handle|install|repair)',
    r'\blicensed\s+(and|&)\s+insured\b',
    r'\b(call|contact|email)\s+(us|me|him|for|at)\b',
    r'\byears?\s+of\s+experience\b',
    r'\bfamily\s+owned\b',
    r'\b(free|experienced|professional|certified)\s+(contractor|builder|specialist)\b',
    r'\b(llc|inc|corp|co\.?|construction|company|services|group)\b',
    r'\b(hire|employment|apply|position|job|w2|1099)\b',
    r'^(looking to hire|we are hiring|job opening|position available)',
    r'(visit|check out|see|look at|follow)\s+(us|our website|our page)',
]

# Homeowner POSITIVE signals
HOMEOWNER_SIGNALS = [
    (r'\b(i\s+)?need\b', 'needs service'),
    (r'\b(my|our|the)\s+(home|house|apartment|kitchen|bathroom|bedroom)\b', 'talking about property'),
    (r'\b(damage|broken|leaking|cracked|water damage|mold)\b', 'damage described'),
    (r'\b(urgent|asap|immediately|soon)\b', 'time-sensitive'),
    (r'\b(budget|price|cost|afford)\b', 'budget mentioned'),
    (r'\b(looking for|seeking|can you|do you)\b', 'actively seeking help'),
]

def load_seen():
    """Load already-seen post IDs"""
    if SEEN_FILE.exists():
        try:
            with open(SEEN_FILE) as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_seen(seen):
    """Save seen post IDs"""
    SEEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SEEN_FILE, 'w') as f:
        json.dump(seen, f, indent=2)

def fetch_page(url, timeout=30):
    """Fetch webpage content"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        req = Request(url, headers=headers)
        with urlopen(req, timeout=timeout) as response:
            return response.read().decode('utf-8', errors='ignore')
    except Exception as e:
        print(f"    Error fetching {url}: {e}")
        return None

def search_craigslist(query):
    """Search Craigslist for a query"""
    # Search in 'household services wanted' section
    url = f"https://chicago.craigslist.org/search/hsh?query={query.replace(' ', '+')}"
    print(f"[*] Searching: '{query}'")
    print(f"    URL: {url}")
    
    html = fetch_page(url)
    if not html:
        print(f"    Failed to fetch page")
        return []
    
    # Extract individual post links
    # Craigslist uses: <a class="posting-title" href="/..../12345.html">Title</a>
    pattern = r'<a[^>]*href="(/[^"]*?/(\d+)\.html)"[^>]*class="[^"]*posting-title[^"]*"[^>]*>([^<]+)</a>'
    matches = re.finditer(pattern, html, re.IGNORECASE)
    
    results = []
    for m in matches:
        rel_url, post_id, title = m.groups()
        post_url = f"https://chicago.craigslist.org{rel_url}"
        results.append({
            'postId': post_id,
            'postUrl': post_url,
            'title': title.strip()[:100]
        })
    
    print(f"    Found {len(results)} listings")
    return results

def validate_post(post_id, post_url, title):
    """Validate if a post is from a homeowner asking for help"""
    html = fetch_page(post_url)
    if not html:
        return None
    
    # Extract post body
    body_match = re.search(
        r'<section id="postingbody"[^>]*>(.+?)</section>',
        html,
        re.DOTALL
    )
    
    full_text = body_match.group(1) if body_match else title
    
    # Clean HTML
    full_text = re.sub(r'<[^>]+>', ' ', full_text)
    full_text = re.sub(r'&[a-z]+;', '', full_text)
    full_text = re.sub(r'\s+', ' ', full_text).strip()
    
    # Combined text for checking
    check_text = f"{title} {full_text}".lower()
    
    result = {
        'postId': post_id,
        'postUrl': post_url,
        'title': title,
        'textSnippet': full_text[:250],
        'valid': True,
        'score': 0,
        'reason': [],
        'signals': [],
        'redFlags': []
    }
    
    # Check RED FLAGS first - if any red flag, it's likely a contractor
    for flag_pattern in RED_FLAGS:
        if re.search(flag_pattern, check_text, re.IGNORECASE):
            result['valid'] = False
            result['redFlags'].append(flag_pattern[:40])
    
    # Check HOMEOWNER SIGNALS
    signal_count = 0
    for signal_pattern, signal_label in HOMEOWNER_SIGNALS:
        if re.search(signal_pattern, check_text, re.IGNORECASE):
            signal_count += 1
            result['signals'].append(signal_label)
            result['score'] += 20
    
    # Final validation logic
    if result['valid']:
        if signal_count == 0:
            result['valid'] = False
            result['reason'].append("No homeowner signals found")
        elif signal_count < 2:
            result['reason'].append(f"Weak signal ({signal_count} indicators)")
        else:
            result['reason'].append("Strong homeowner lead")
    else:
        result['reason'].append(f"Contractor red flags: {len(result['redFlags'])} detected")
    
    return result

def log_lead_to_tracker(lead):
    """Log valid lead to Excel tracker"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(TRACKER_PATH)
        # Try both sheet names
        sheet_name = None
        for name in ['Facebook Leads', 'Lead Tracker']:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"    [!] No suitable sheet found in tracker")
            return False
        
        ws = wb[sheet_name]
        
        # Find next empty row (start at row 3)
        next_row = 3
        for row in range(3, 1000):
            if ws[f'B{row}'].value is None:
                next_row = row
                break
        
        # Log the lead
        ws[f'B{next_row}'] = datetime.now()
        ws[f'C{next_row}'] = 'Craigslist'
        ws[f'D{next_row}'] = 'Home Services'
        ws[f'I{next_row}'] = lead['title'][:100]
        ws[f'K{next_row}'] = 'Hot'
        ws[f'L{next_row}'] = 'New'
        ws[f'P{next_row}'] = lead['postUrl']
        
        wb.save(TRACKER_PATH)
        print(f"    [+] Added to tracker")
        return True
    except Exception as e:
        print(f"    [!] Tracker error: {e}")
        return False

def main():
    seen = load_seen()
    valid_leads = []
    rejected = []
    checked = 0
    
    print("="*60)
    print("STRICT HOMEOWNER LEAD HUNTER - CRAIGSLIST")
    print("="*60 + "\n")
    
    for query in SEARCH_PHRASES:
        results = search_craigslist(query)
        time.sleep(1.5)
        
        for result in results:
            post_id = result['postId']
            
            # Skip if already seen
            if post_id in seen:
                continue
            
            checked += 1
            print(f"  Validating: {result['title'][:50]}...")
            
            validation = validate_post(
                post_id,
                result['postUrl'],
                result['title']
            )
            
            if validation:
                if validation['valid']:
                    print(f"    [ACCEPT] Homeowner lead")
                    print(f"    Signals: {', '.join(validation['signals'])}")
                    print(f"    Score: {validation['score']}")
                    valid_leads.append(validation)
                    seen[post_id] = datetime.now().isoformat()
                    log_lead_to_tracker(validation)
                else:
                    print(f"    [REJECT] {validation['reason'][0]}")
                    rejected.append(validation)
            
            time.sleep(0.3)
    
    # Save tracking
    save_seen(seen)
    
    # Generate report
    quality = (len(valid_leads) / checked * 100) if checked > 0 else 0
    
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'totalChecked': checked,
            'validLeadsFound': len(valid_leads),
            'rejectedLeads': len(rejected),
            'qualityScorePercent': round(quality, 1)
        },
        'validLeads': valid_leads,
        'rejectedSample': rejected[:5]
    }
    
    # Save report
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, indent=2)
    
    # Print summary
    print("\n" + "="*60)
    print("HUNT SUMMARY")
    print("="*60)
    print(f"Total Posts Checked: {checked}")
    print(f"Valid Homeowner Leads: {len(valid_leads)}")
    print(f"Rejected (Contractor): {len(rejected)}")
    print(f"Quality Score: {quality}%")
    print(f"Report: {REPORT_FILE}")
    
    if valid_leads:
        print("\nVALID LEADS FOUND:")
        for i, lead in enumerate(valid_leads, 1):
            print(f"\n{i}. {lead['title']}")
            print(f"   URL: {lead['postUrl']}")
            print(f"   Signals: {', '.join(lead['signals'])}")
    else:
        print("\nNo valid homeowner leads found in this run.")
    
    print("\n" + "="*60)
    return json.dumps(report, indent=2, default=str)

if __name__ == '__main__':
    output = main()
    sys.exit(0)
